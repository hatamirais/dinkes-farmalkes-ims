from decimal import Decimal, InvalidOperation
import unicodedata

from django.core.exceptions import ValidationError
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.core.decimal_validation import validate_finite_decimal
from apps.core.xlsx_exports import escape_xlsx_formula
from apps.lplpo.models import get_previous_lplpo, is_january_bootstrap_period


WORKSHEET_TITLE = "LPLPO Offline Entry"
HEADER_ROW = 11
DATA_START_ROW = 12
EDITABLE_COLUMNS_LABEL = (
    "stock_awal, penerimaan, harga_satuan, stock_gudang_puskesmas, "
    "waktu_kosong, permintaan_jumlah, permintaan_alasan"
)
HEADER_COLUMNS = [
    "item_code",
    "item_name",
    "unit",
    "category",
    "stock_awal",
    "penerimaan",
    "harga_satuan",
    "pemakaian",
    "persediaan",
    "stock_keseluruhan",
    "stock_optimum",
    "jumlah_kebutuhan",
    "stock_gudang_puskesmas",
    "waktu_kosong",
    "permintaan_jumlah",
    "permintaan_alasan",
]
MAX_XLSX_SIZE_BYTES = 2 * 1024 * 1024

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
META_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
INSTRUCTION_FILL = PatternFill(
    start_color="FFF3CD",
    end_color="FFF3CD",
    fill_type="solid",
)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
NUMBER_FORMAT = "#,##0"
DECIMAL_FORMAT = "#,##0.00"


def _cell_value(value):
    return escape_xlsx_formula(value)


def _normalize_text_value(value, *, field_label, max_length=None, allow_empty=True):
    if value in (None, ""):
        if allow_empty:
            return ""
        raise ValidationError(f"{field_label} wajib diisi.")

    normalized = unicodedata.normalize("NFC", str(value)).strip()
    if not normalized:
        if allow_empty:
            return ""
        raise ValidationError(f"{field_label} wajib diisi.")
    if "\x00" in normalized:
        raise ValidationError(f"{field_label} tidak boleh mengandung null byte.")
    if max_length is not None and len(normalized) > max_length:
        raise ValidationError(
            f"{field_label} tidak boleh lebih dari {max_length} karakter."
        )
    return normalized


def _parse_decimal_from_cell(value, *, field_label):
    if value in (None, ""):
        return Decimal("0")

    if isinstance(value, str):
        normalized = _normalize_text_value(
            value,
            field_label=field_label,
            max_length=100,
            allow_empty=False,
        ).replace(" ", "")
        if "," in normalized and "." in normalized:
            normalized = normalized.replace(".", "").replace(",", ".")
        elif "," in normalized:
            normalized = normalized.replace(",", ".")
        raw_value = normalized
    else:
        raw_value = str(value).strip()

    try:
        decimal_value = Decimal(raw_value)
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValidationError(f"{field_label} tidak valid.") from exc

    return validate_finite_decimal(decimal_value, field_label=field_label)


def _parse_integer_from_cell(value, *, field_label, allow_negative=False):
    decimal_value = _parse_decimal_from_cell(value, field_label=field_label)
    if not allow_negative and decimal_value < 0:
        raise ValidationError(f"{field_label} tidak boleh negatif.")
    if decimal_value != decimal_value.to_integral_value():
        raise ValidationError(f"{field_label} harus berupa bilangan bulat.")
    return int(decimal_value)


def _validate_no_formula(cell, *, field_label):
    if cell.data_type == "f":
        raise ValidationError(f"{field_label} tidak boleh berupa formula Excel.")


def export_lplpo_workbook(lplpo_obj):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = WORKSHEET_TITLE
    worksheet.freeze_panes = f"A{DATA_START_ROW}"

    metadata_rows = [
        ("document_number", lplpo_obj.document_number),
        ("facility_code", lplpo_obj.facility.code),
        ("facility_name", lplpo_obj.facility.name),
        ("bulan", lplpo_obj.bulan),
        ("tahun", lplpo_obj.tahun),
        ("status", lplpo_obj.status),
        ("editable_columns", EDITABLE_COLUMNS_LABEL),
    ]

    worksheet["A1"] = _cell_value("LPLPO Offline Entry")
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet.merge_cells("A1:D1")

    for row_num, (label, value) in enumerate(metadata_rows, start=2):
        label_cell = worksheet.cell(row=row_num, column=1, value=_cell_value(label))
        value_cell = worksheet.cell(row=row_num, column=2, value=_cell_value(value))
        label_cell.font = HEADER_FONT
        label_cell.fill = META_FILL
        label_cell.border = THIN_BORDER
        value_cell.border = THIN_BORDER

    worksheet.merge_cells(start_row=9, start_column=1, end_row=9, end_column=16)
    instruction_cell = worksheet.cell(
        row=9,
        column=1,
        value=_cell_value(
            "Edit only these columns: stock_awal, penerimaan, harga_satuan, "
            "stock_gudang_puskesmas, waktu_kosong, permintaan_jumlah, "
            "permintaan_alasan. Do not change metadata, item_code, or computed columns."
        ),
    )
    instruction_cell.fill = INSTRUCTION_FILL
    instruction_cell.border = THIN_BORDER
    instruction_cell.alignment = Alignment(wrap_text=True)

    for column_index, header in enumerate(HEADER_COLUMNS, start=1):
        header_cell = worksheet.cell(
            row=HEADER_ROW,
            column=column_index,
            value=_cell_value(header),
        )
        header_cell.font = HEADER_FONT
        header_cell.fill = HEADER_FILL
        header_cell.border = THIN_BORDER
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, line in enumerate(
        lplpo_obj.items.select_related("item", "item__satuan", "item__kategori").order_by(
            "item__kategori__sort_order", "item__nama_barang"
        ),
        start=DATA_START_ROW,
    ):
        row_values = [
            line.item.kode_barang,
            line.item.nama_barang,
            line.item.satuan.code,
            line.item.kategori.name if line.item.kategori else "",
            line.stock_awal,
            line.penerimaan,
            line.harga_satuan,
            line.pemakaian,
            line.persediaan,
            line.stock_keseluruhan,
            line.stock_optimum,
            line.jumlah_kebutuhan,
            line.stock_gudang_puskesmas,
            line.waktu_kosong,
            line.permintaan_jumlah,
            line.permintaan_alasan,
        ]
        for column_index, value in enumerate(row_values, start=1):
            cell = worksheet.cell(
                row=row_index,
                column=column_index,
                value=_cell_value(value),
            )
            cell.border = THIN_BORDER
            if column_index in {5, 6, 8, 9, 10, 13, 14, 15}:
                cell.number_format = NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="right")
            elif column_index in {7, 11, 12}:
                cell.number_format = DECIMAL_FORMAT
                cell.alignment = Alignment(horizontal="right")

    column_widths = [18, 34, 10, 20, 12, 12, 14, 12, 12, 16, 14, 16, 14, 12, 16, 36]
    for column_index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response[
        "Content-Disposition"
    ] = f'attachment; filename="LPLPO_{lplpo_obj.document_number}_offline.xlsx"'
    workbook.save(response)
    return response


def apply_lplpo_workbook_import(*, uploaded_file, lplpo_obj):
    workbook = load_workbook(uploaded_file, data_only=False)
    try:
        worksheet = workbook[WORKSHEET_TITLE]
    except KeyError as exc:
        workbook.close()
        raise ValidationError("Sheet XLSX tidak valid.") from exc

    metadata_map = {
        "document_number": lplpo_obj.document_number,
        "facility_code": lplpo_obj.facility.code,
        "facility_name": lplpo_obj.facility.name,
        "bulan": lplpo_obj.bulan,
        "tahun": lplpo_obj.tahun,
        "status": lplpo_obj.status,
    }

    for row_num, (key, expected_value) in enumerate(metadata_map.items(), start=2):
        label_cell = worksheet.cell(row=row_num, column=1)
        value_cell = worksheet.cell(row=row_num, column=2)
        _validate_no_formula(label_cell, field_label=f"Metadata {key}")
        _validate_no_formula(value_cell, field_label=f"Metadata {key}")

        if str(label_cell.value or "").strip() != key:
            workbook.close()
            raise ValidationError("Template XLSX tidak sesuai.")
        if str(value_cell.value or "").strip() != str(expected_value):
            workbook.close()
            raise ValidationError("Metadata XLSX tidak cocok dengan dokumen LPLPO ini.")

    header_values = [
        str(worksheet.cell(row=HEADER_ROW, column=column_index).value or "").strip()
        for column_index in range(1, len(HEADER_COLUMNS) + 1)
    ]
    if header_values != HEADER_COLUMNS:
        workbook.close()
        raise ValidationError("Header XLSX tidak valid.")

    line_by_item_code = {
        line.item.kode_barang: line
        for line in lplpo_obj.items.select_related("item").order_by("item__kode_barang")
    }
    stock_awal_locked = (
        get_previous_lplpo(lplpo_obj.facility, lplpo_obj.bulan, lplpo_obj.tahun)
        is not None
        and not is_january_bootstrap_period(lplpo_obj.bulan, lplpo_obj.tahun)
    )
    seen_item_codes = set()
    updated_lines = []
    non_empty_rows = []

    for row_num in range(DATA_START_ROW, worksheet.max_row + 1):
        row_cells = [worksheet.cell(row=row_num, column=col) for col in range(1, 17)]
        if not any(cell.value not in (None, "") for cell in row_cells):
            continue
        non_empty_rows.append(row_num)

        for index, cell in enumerate(row_cells, start=1):
            if index in {1, 5, 6, 7, 13, 14, 15, 16}:
                _validate_no_formula(cell, field_label=f"Baris {row_num}")

        item_code = _normalize_text_value(
            row_cells[0].value,
            field_label=f"Baris {row_num} item_code",
            max_length=100,
            allow_empty=False,
        )
        if item_code in seen_item_codes:
            workbook.close()
            raise ValidationError(f"Baris {row_num}: item_code duplikat '{item_code}'.")
        seen_item_codes.add(item_code)

        line = line_by_item_code.get(item_code)
        if line is None:
            workbook.close()
            raise ValidationError(
                f"Baris {row_num}: item_code '{item_code}' tidak cocok dengan dokumen ini."
            )
        if not line.item.is_active:
            workbook.close()
            raise ValidationError(
                f"Baris {row_num}: item '{item_code}' sudah tidak aktif."
            )

        imported_stock_awal = _parse_integer_from_cell(
            row_cells[4].value,
            field_label=f"Baris {row_num} stock_awal",
            allow_negative=True,
        )
        if stock_awal_locked and imported_stock_awal != line.stock_awal:
            workbook.close()
            raise ValidationError(
                f"Baris {row_num}: stock_awal terkunci mengikuti sisa stok bulan sebelumnya."
            )
        line.stock_awal = imported_stock_awal
        line.penerimaan = _parse_integer_from_cell(
            row_cells[5].value,
            field_label=f"Baris {row_num} penerimaan",
        )
        harga_satuan = _parse_decimal_from_cell(
            row_cells[6].value,
            field_label=f"Baris {row_num} harga_satuan",
        )
        if harga_satuan < 0:
            workbook.close()
            raise ValidationError(f"Baris {row_num}: harga_satuan tidak boleh negatif.")
        line.harga_satuan = harga_satuan.quantize(Decimal("0.01"))
        line.stock_gudang_puskesmas = _parse_integer_from_cell(
            row_cells[12].value,
            field_label=f"Baris {row_num} stock_gudang_puskesmas",
        )
        line.waktu_kosong = _parse_integer_from_cell(
            row_cells[13].value,
            field_label=f"Baris {row_num} waktu_kosong",
        )
        line.permintaan_jumlah = _parse_integer_from_cell(
            row_cells[14].value,
            field_label=f"Baris {row_num} permintaan_jumlah",
        )
        line.permintaan_alasan = _normalize_text_value(
            row_cells[15].value,
            field_label=f"Baris {row_num} permintaan_alasan",
            max_length=1000,
        )
        line.penerimaan_auto_filled = False
        line.compute_fields()
        updated_lines.append(line)

    workbook.close()

    if len(non_empty_rows) != len(line_by_item_code):
        missing_count = len(line_by_item_code) - len(non_empty_rows)
        if missing_count > 0:
            raise ValidationError(
                "Jumlah baris item pada XLSX tidak lengkap untuk dokumen LPLPO ini."
            )
        raise ValidationError(
            "Jumlah baris item pada XLSX melebihi item dokumen LPLPO ini."
        )

    if seen_item_codes != set(line_by_item_code):
        raise ValidationError("Item pada XLSX harus cocok persis dengan dokumen LPLPO.")

    return updated_lines
