from io import BytesIO
from decimal import Decimal

from django.contrib.messages import get_messages
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from django.urls import reverse
from openpyxl import load_workbook

from apps.lplpo.models import LPLPO, LPLPOItem
from apps.lplpo.tests import LPLPOTestCase


@override_settings(LPLPO_IMPORT_RATE_LIMIT="1000/m", RATELIMIT_FAIL_OPEN=False)
class LPLPOOfflineXlsxTests(LPLPOTestCase):
    def create_lplpo_with_items(self, *, status=LPLPO.Status.DRAFT):
        lplpo = self.create_lplpo(status=status, created_by=self.puskesmas_user)
        line_a = LPLPOItem.objects.create(
            lplpo=lplpo,
            item=self.item_a,
            stock_awal=10,
            penerimaan=4,
            harga_satuan=Decimal("1250.00"),
            pemakaian=3,
            stock_gudang_puskesmas=2,
            waktu_kosong=1,
            permintaan_jumlah=5,
            permintaan_alasan="=offline-note",
            pemberian_jumlah=7,
        )
        line_b = LPLPOItem.objects.create(
            lplpo=lplpo,
            item=self.item_b,
            stock_awal=8,
            penerimaan=2,
            harga_satuan=Decimal("2000.00"),
            pemakaian=6,
            stock_gudang_puskesmas=1,
            waktu_kosong=0,
            permintaan_jumlah=4,
            permintaan_alasan="Catatan biasa",
        )
        return lplpo, line_a, line_b

    @staticmethod
    def _workbook_upload(workbook):
        file_obj = BytesIO()
        workbook.save(file_obj)
        file_obj.seek(0)
        return SimpleUploadedFile(
            "lplpo-offline.xlsx",
            file_obj.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _export_workbook(self, lplpo):
        response = self.client.get(reverse("lplpo:lplpo_export_xlsx", args=[lplpo.pk]))
        self.assertEqual(response.status_code, 200)
        return load_workbook(BytesIO(response.content))

    @staticmethod
    def _find_row_by_item_code(worksheet, item_code):
        for row_num in range(12, worksheet.max_row + 1):
            if worksheet.cell(row=row_num, column=1).value == item_code:
                return row_num
        raise AssertionError(f"item_code {item_code} not found in exported workbook")

    def test_export_xlsx_returns_workbook_with_metadata_and_rows(self):
        lplpo, _, _ = self.create_lplpo_with_items()
        self.client.force_login(self.puskesmas_user)

        response = self.client.get(reverse("lplpo:lplpo_export_xlsx", args=[lplpo.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook["LPLPO Offline Entry"]
        row_a = self._find_row_by_item_code(worksheet, self.item_a.kode_barang)
        row_b = self._find_row_by_item_code(worksheet, self.item_b.kode_barang)
        self.assertEqual(worksheet["A2"].value, "document_number")
        self.assertEqual(worksheet["B2"].value, lplpo.document_number)
        self.assertEqual(worksheet["B3"].value, self.facility.code)
        self.assertEqual(worksheet["B5"].value, lplpo.bulan)
        self.assertEqual(worksheet["A11"].value, "item_code")
        self.assertEqual(worksheet.cell(row=row_a, column=1).value, self.item_a.kode_barang)
        self.assertEqual(worksheet.cell(row=row_b, column=1).value, self.item_b.kode_barang)

    def test_export_xlsx_escapes_formula_like_text(self):
        lplpo, _, _ = self.create_lplpo_with_items()
        self.client.force_login(self.puskesmas_user)

        workbook = self._export_workbook(lplpo)
        worksheet = workbook["LPLPO Offline Entry"]
        row_a = self._find_row_by_item_code(worksheet, self.item_a.kode_barang)

        self.assertEqual(worksheet.cell(row=row_a, column=16).value, "'=offline-note")

    def test_import_xlsx_updates_editable_fields_and_preserves_server_owned_values(self):
        lplpo, line_a, line_b = self.create_lplpo_with_items()
        self.client.force_login(self.puskesmas_user)
        workbook = self._export_workbook(lplpo)
        worksheet = workbook["LPLPO Offline Entry"]
        row_a = self._find_row_by_item_code(worksheet, self.item_a.kode_barang)
        row_b = self._find_row_by_item_code(worksheet, self.item_b.kode_barang)

        worksheet.cell(row=row_a, column=5).value = 14
        worksheet.cell(row=row_a, column=6).value = 9
        worksheet.cell(row=row_a, column=7).value = 1500
        worksheet.cell(row=row_a, column=8).value = 999
        worksheet.cell(row=row_a, column=13).value = 5
        worksheet.cell(row=row_a, column=14).value = 2
        worksheet.cell(row=row_a, column=15).value = 8
        worksheet.cell(row=row_a, column=16).value = "Revisi offline"
        worksheet.cell(row=row_b, column=5).value = 1
        worksheet.cell(row=row_b, column=6).value = 3
        worksheet.cell(row=row_b, column=7).value = 2500
        worksheet.cell(row=row_b, column=13).value = 4
        worksheet.cell(row=row_b, column=14).value = 1
        worksheet.cell(row=row_b, column=15).value = 9

        with self.assertLogs("security", level="INFO") as logs:
            response = self.client.post(
                reverse("lplpo:lplpo_import_xlsx", args=[lplpo.pk]),
                {"xlsx_file": self._workbook_upload(workbook)},
                follow=True,
            )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            any("lplpo_xlsx_import_succeeded" in message for message in logs.output)
        )

        line_a.refresh_from_db()
        line_b.refresh_from_db()
        self.assertEqual(line_a.stock_awal, 14)
        self.assertEqual(line_a.penerimaan, 9)
        self.assertEqual(line_a.harga_satuan, Decimal("1500.00"))
        self.assertEqual(line_a.pemakaian, 3)
        self.assertEqual(line_a.stock_gudang_puskesmas, 5)
        self.assertEqual(line_a.waktu_kosong, 2)
        self.assertEqual(line_a.permintaan_jumlah, 8)
        self.assertEqual(line_a.permintaan_alasan, "Revisi offline")
        self.assertEqual(line_a.pemberian_jumlah, 7)
        self.assertFalse(line_a.penerimaan_auto_filled)
        self.assertEqual(line_a.persediaan, 23)
        self.assertEqual(line_a.stock_keseluruhan, 20)

        self.assertEqual(line_b.stock_awal, 1)
        self.assertEqual(line_b.penerimaan, 3)
        self.assertEqual(line_b.harga_satuan, Decimal("2500.00"))
        self.assertEqual(line_b.pemakaian, 6)
        self.assertEqual(line_b.stock_gudang_puskesmas, 4)
        self.assertEqual(line_b.waktu_kosong, 1)
        self.assertEqual(line_b.permintaan_jumlah, 9)

    def test_import_xlsx_rejects_stock_awal_change_when_carry_forward_is_locked(self):
        previous = self.create_lplpo(
            bulan=1,
            tahun=2026,
            status=LPLPO.Status.CLOSED,
        )
        LPLPOItem.objects.create(
            lplpo=previous,
            item=self.item_a,
            stock_awal=5,
            penerimaan=1,
            harga_satuan=Decimal("1000.00"),
            pemakaian=2,
            stock_gudang_puskesmas=1,
            waktu_kosong=0,
            permintaan_jumlah=1,
        )
        lplpo, line_a, _ = self.create_lplpo_with_items()
        self.client.force_login(self.puskesmas_user)
        workbook = self._export_workbook(lplpo)
        worksheet = workbook["LPLPO Offline Entry"]
        row_a = self._find_row_by_item_code(worksheet, self.item_a.kode_barang)
        worksheet.cell(row=row_a, column=5).value = 99

        response = self.client.post(
            reverse("lplpo:lplpo_import_xlsx", args=[lplpo.pk]),
            {"xlsx_file": self._workbook_upload(workbook)},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            f"Baris {row_a}: stock_awal terkunci mengikuti sisa stok bulan sebelumnya.",
            [str(message) for message in get_messages(response.wsgi_request)],
        )
        line_a.refresh_from_db()
        self.assertEqual(line_a.stock_awal, 10)

    def test_import_xlsx_allows_stock_awal_change_for_january_bootstrap(self):
        lplpo, line_a, _ = self.create_lplpo_with_items()
        lplpo.bulan = 1
        lplpo.tahun = 2026
        lplpo.document_number = ""
        lplpo.save(update_fields=["bulan", "tahun", "document_number", "updated_at"])
        self.client.force_login(self.puskesmas_user)
        workbook = self._export_workbook(lplpo)
        worksheet = workbook["LPLPO Offline Entry"]
        row_a = self._find_row_by_item_code(worksheet, self.item_a.kode_barang)
        worksheet.cell(row=row_a, column=5).value = 21

        response = self.client.post(
            reverse("lplpo:lplpo_import_xlsx", args=[lplpo.pk]),
            {"xlsx_file": self._workbook_upload(workbook)},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        line_a.refresh_from_db()
        self.assertEqual(line_a.stock_awal, 21)

    def test_import_xlsx_rejects_metadata_mismatch(self):
        lplpo, line_a, _ = self.create_lplpo_with_items()
        self.client.force_login(self.puskesmas_user)
        workbook = self._export_workbook(lplpo)
        worksheet = workbook["LPLPO Offline Entry"]
        worksheet["B2"] = "WRONG-DOC"

        with self.assertLogs("security", level="WARNING") as logs:
            response = self.client.post(
                reverse("lplpo:lplpo_import_xlsx", args=[lplpo.pk]),
                {"xlsx_file": self._workbook_upload(workbook)},
                follow=True,
            )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            any("lplpo_xlsx_import_failed" in message for message in logs.output)
        )
        self.assertIn(
            "Metadata XLSX tidak cocok dengan dokumen LPLPO ini.",
            [str(message) for message in get_messages(response.wsgi_request)],
        )
        line_a.refresh_from_db()
        self.assertEqual(line_a.stock_awal, 10)

    def test_import_xlsx_rejects_formula_in_editable_cell(self):
        lplpo, line_a, _ = self.create_lplpo_with_items()
        self.client.force_login(self.puskesmas_user)
        workbook = self._export_workbook(lplpo)
        worksheet = workbook["LPLPO Offline Entry"]
        row_a = self._find_row_by_item_code(worksheet, self.item_a.kode_barang)
        worksheet.cell(row=row_a, column=5).value = "=1+1"

        response = self.client.post(
            reverse("lplpo:lplpo_import_xlsx", args=[lplpo.pk]),
            {"xlsx_file": self._workbook_upload(workbook)},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            f"Baris {row_a} tidak boleh berupa formula Excel.",
            [str(message) for message in get_messages(response.wsgi_request)],
        )
        line_a.refresh_from_db()
        self.assertEqual(line_a.stock_awal, 10)

    def test_import_xlsx_rejects_missing_item_row(self):
        lplpo, line_a, _ = self.create_lplpo_with_items()
        self.client.force_login(self.puskesmas_user)
        workbook = self._export_workbook(lplpo)
        worksheet = workbook["LPLPO Offline Entry"]
        row_b = self._find_row_by_item_code(worksheet, self.item_b.kode_barang)
        for column_index in range(1, 17):
            worksheet.cell(row=row_b, column=column_index).value = None

        response = self.client.post(
            reverse("lplpo:lplpo_import_xlsx", args=[lplpo.pk]),
            {"xlsx_file": self._workbook_upload(workbook)},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "Jumlah baris item pada XLSX tidak lengkap untuk dokumen LPLPO ini.",
            [str(message) for message in get_messages(response.wsgi_request)],
        )
        line_a.refresh_from_db()
        self.assertEqual(line_a.stock_awal, 10)

    def test_non_owner_facility_cannot_import_xlsx(self):
        lplpo, _, _ = self.create_lplpo_with_items()
        self.client.force_login(self.puskesmas_user)
        workbook = self._export_workbook(lplpo)
        self.client.force_login(self.other_puskesmas_user)

        response = self.client.post(
            reverse("lplpo:lplpo_import_xlsx", args=[lplpo.pk]),
            {"xlsx_file": self._workbook_upload(workbook)},
        )

        self.assertEqual(response.status_code, 403)

    def test_non_puskesmas_staff_cannot_export_or_import_xlsx(self):
        lplpo, _, _ = self.create_lplpo_with_items()
        self.client.force_login(self.gudang_user)

        export_response = self.client.get(
            reverse("lplpo:lplpo_export_xlsx", args=[lplpo.pk])
        )
        import_response = self.client.get(
            reverse("lplpo:lplpo_import_xlsx", args=[lplpo.pk])
        )

        self.assertEqual(export_response.status_code, 403)
        self.assertEqual(import_response.status_code, 403)

    def test_submitted_lplpo_cannot_export_or_import_xlsx(self):
        lplpo, _, _ = self.create_lplpo_with_items(status=LPLPO.Status.SUBMITTED)
        self.client.force_login(self.puskesmas_user)

        export_response = self.client.get(
            reverse("lplpo:lplpo_export_xlsx", args=[lplpo.pk]),
            follow=True,
        )
        import_response = self.client.get(
            reverse("lplpo:lplpo_import_xlsx", args=[lplpo.pk]),
            follow=True,
        )

        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(import_response.status_code, 200)
        self.assertContains(
            export_response,
            "Hanya LPLPO berstatus Draft atau Ditolak yang dapat diekspor untuk pengisian offline.",
        )
        self.assertContains(
            import_response,
            "Hanya LPLPO berstatus Draft atau Ditolak yang dapat diimpor dari XLSX.",
        )

    def test_import_xlsx_form_rejects_non_xlsx_upload(self):
        lplpo, _, _ = self.create_lplpo_with_items()
        self.client.force_login(self.puskesmas_user)

        response = self.client.post(
            reverse("lplpo:lplpo_import_xlsx", args=[lplpo.pk]),
            {
                "xlsx_file": SimpleUploadedFile(
                    "not-xlsx.csv",
                    b"item_code\nA\n",
                    content_type="text/csv",
                )
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "File XLSX harus memakai ekstensi yang diizinkan: .xlsx.")

    def test_import_xlsx_accepts_valid_workbook_with_octet_stream_content_type(self):
        lplpo, line_a, _ = self.create_lplpo_with_items()
        self.client.force_login(self.puskesmas_user)
        workbook = self._export_workbook(lplpo)
        file_obj = BytesIO()
        workbook.save(file_obj)
        file_obj.seek(0)

        response = self.client.post(
            reverse("lplpo:lplpo_import_xlsx", args=[lplpo.pk]),
            {
                "xlsx_file": SimpleUploadedFile(
                    "lplpo-offline.xlsx",
                    file_obj.read(),
                    content_type="application/octet-stream",
                )
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        line_a.refresh_from_db()
        self.assertEqual(line_a.stock_awal, 10)

    @override_settings(LPLPO_IMPORT_RATE_LIMIT="1/m", RATELIMIT_FAIL_OPEN=False)
    def test_import_xlsx_is_rate_limited(self):
        lplpo, _, _ = self.create_lplpo_with_items()
        self.client.force_login(self.puskesmas_user)
        workbook = self._export_workbook(lplpo)
        upload = self._workbook_upload(workbook)

        first_response = self.client.post(
            reverse("lplpo:lplpo_import_xlsx", args=[lplpo.pk]),
            {"xlsx_file": upload},
        )
        self.assertEqual(first_response.status_code, 302)

        second_workbook = self._export_workbook(lplpo)
        second_response = self.client.post(
            reverse("lplpo:lplpo_import_xlsx", args=[lplpo.pk]),
            {"xlsx_file": self._workbook_upload(second_workbook)},
        )

        self.assertEqual(second_response.status_code, 429)
