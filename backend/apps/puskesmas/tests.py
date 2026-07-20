from decimal import Decimal
from io import BytesIO
from datetime import timedelta
from unittest.mock import patch

from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ValidationError
from django.db import IntegrityError, connection
from django.db.migrations.executor import MigrationExecutor
from django.test import TestCase, TransactionTestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.core.tests.mixins import SecureClientDefaultsMixin
from apps.distribution.models import Distribution, DistributionItem
from apps.items.models import Category, Facility, Item, Unit
from apps.puskesmas.exports import export_puskesmas_penerimaan_excel
from apps.puskesmas.forms import (
	PuskesmasConsumptionMatrixForm,
	PuskesmasRequestForm,
	PuskesmasRequestItemForm,
	PuskesmasSBBKForm,
	PuskesmasSBBKItemForm,
	PuskesmasSubunitForm,
)
from apps.puskesmas.models import (
	PuskesmasConsumption,
	PuskesmasConsumptionEntry,
	PuskesmasRequest,
	PuskesmasRequestItem,
	PuskesmasSBBK,
	PuskesmasSBBKItem,
	PuskesmasSubunit,
)
from apps.puskesmas.services import assert_consumption_month_mutable
from apps.puskesmas.views import _get_consumption_subunits
from apps.users.access import ensure_default_module_access
from apps.users.models import ModuleAccess, User
from apps.lplpo.models import LPLPO, LPLPOItem


class PuskesmasRequestFormTests(TestCase):
	def setUp(self):
		self.unit = Unit.objects.create(code="TAB", name="Tablet")
		self.category = Category.objects.create(code="OBT", name="Obat", sort_order=1)
		self.facility = Facility.objects.create(
			code="PKM-01",
			name="Puskesmas Satu",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.other_facility = Facility.objects.create(
			code="PKM-02",
			name="Puskesmas Dua",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.item = Item.objects.create(
			nama_barang="Haloperidol 5 mg",
			satuan=self.unit,
			kategori=self.category,
			is_active=True,
		)
		self.user = User.objects.create_user(
			username="operator-puskesmas",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=self.facility,
		)

	def test_operator_form_hides_facility_and_uses_account_facility(self):
		form = PuskesmasRequestForm(user=self.user)

		self.assertEqual(form.fields["facility"].widget.__class__.__name__, "HiddenInput")
		self.assertEqual(form.fields["facility"].initial, self.facility.pk)

	def test_operator_form_ignores_posted_other_facility(self):
		form = PuskesmasRequestForm(
			data={
				"document_number": "",
				"facility": self.other_facility.pk,
				"request_date": "2026-04-02",
				"program": "",
				"notes": "",
			},
			user=self.user,
		)

		self.assertTrue(form.is_valid())
		self.assertEqual(form.cleaned_data["facility"], self.facility)

	def test_item_label_shows_name_only(self):
		self.item.nama_barang = "Haloperidol 5 mg [P]"
		self.item.save(update_fields=["nama_barang", "updated_at"])

		form = PuskesmasRequestItemForm()

		self.assertEqual(form.fields["item"].label_from_instance(self.item), "Haloperidol 5 mg")


class PuskesmasStockCurrentViewTests(SecureClientDefaultsMixin, TestCase):
	def setUp(self):
		super().setUp()
		self.unit = Unit.objects.create(code="TAB-STK", name="Tablet")
		self.category = Category.objects.create(
			code="OBT-STK",
			name="Obat",
			sort_order=1,
		)
		self.facility = Facility.objects.create(
			code="PKM-STK-01",
			name="Puskesmas Stok Satu",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.other_facility = Facility.objects.create(
			code="PKM-STK-02",
			name="Puskesmas Stok Dua",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.item = Item.objects.create(
			kode_barang="STK-001",
			nama_barang="Paracetamol 500 mg",
			satuan=self.unit,
			kategori=self.category,
			minimum_stock=Decimal("0"),
		)
		self.matching_item = Item.objects.create(
			kode_barang="STK-002",
			nama_barang="Vitamin B Kompleks",
			satuan=self.unit,
			kategori=self.category,
			minimum_stock=Decimal("0"),
		)
		self.other_item = Item.objects.create(
			kode_barang="STK-003",
			nama_barang="Amoxicillin",
			satuan=self.unit,
			kategori=self.category,
			minimum_stock=Decimal("0"),
		)
		self.user = User.objects.create_user(
			username="puskesmas-stock-current",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=self.facility,
		)
		self.internal_user = User.objects.create_user(
			username="gudang-stock-current",
			password="TestPassword123!",
			role=User.Role.GUDANG,
		)
		self.year = timezone.localdate().year

		self.latest_lplpo = LPLPO.objects.create(
			facility=self.facility,
			bulan=5,
			tahun=self.year,
			status=LPLPO.Status.SUBMITTED,
			created_by=self.user,
		)
		LPLPOItem.objects.create(
			lplpo=self.latest_lplpo,
			item=self.item,
			stock_awal=20,
			penerimaan=5,
			pemakaian=10,
			stock_gudang_puskesmas=12,
		)
		LPLPOItem.objects.create(
			lplpo=self.latest_lplpo,
			item=self.matching_item,
			stock_awal=10,
			penerimaan=0,
			pemakaian=2,
			stock_gudang_puskesmas=8,
		)
		rejected_lplpo = LPLPO.objects.create(
			facility=self.facility,
			bulan=6,
			tahun=self.year,
			status=LPLPO.Status.REJECTED_PUSKESMAS,
			created_by=self.user,
		)
		LPLPOItem.objects.create(
			lplpo=rejected_lplpo,
			item=self.item,
			stock_awal=999,
			penerimaan=0,
			pemakaian=0,
			stock_gudang_puskesmas=999,
		)
		other_lplpo = LPLPO.objects.create(
			facility=self.other_facility,
			bulan=7,
			tahun=self.year,
			status=LPLPO.Status.CLOSED,
			created_by=self.user,
		)
		LPLPOItem.objects.create(
			lplpo=other_lplpo,
			item=self.other_item,
			stock_awal=50,
			penerimaan=0,
			pemakaian=0,
			stock_gudang_puskesmas=50,
		)

	def test_puskesmas_user_sees_latest_usable_lplpo_for_own_facility(self):
		self.client.force_login(self.user)

		response = self.client.get(reverse("puskesmas:stock_current"))

		self.assertEqual(response.status_code, 200)
		self.assertTemplateUsed(response, "puskesmas/stock_current.html")
		self.assertEqual(response.context["facility"], self.facility)
		self.assertEqual(response.context["latest_lplpo"], self.latest_lplpo)
		self.assertContains(response, "Paracetamol 500 mg")
		self.assertContains(response, "Stok Digital")
		self.assertNotContains(response, "Amoxicillin")
		row = next(row for row in response.context["rows"] if row["kode_barang"] == "STK-001")
		self.assertEqual(row["digital_stock"], 15)
		self.assertEqual(row["physical_stock"], 12)
		self.assertEqual(row["difference"], -3)
		self.assertTrue(row["has_mismatch"])

	def test_unlinked_puskesmas_user_receives_403(self):
		unlinked = User.objects.create_user(
			username="puskesmas-stock-unlinked",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
		)
		self.client.force_login(unlinked)

		response = self.client.get(reverse("puskesmas:stock_current"))

		self.assertEqual(response.status_code, 403)

	def test_non_puskesmas_user_cannot_access_page(self):
		ModuleAccess.objects.update_or_create(
			user=self.internal_user,
			module=ModuleAccess.Module.PUSKESMAS,
			defaults={"scope": ModuleAccess.Scope.MANAGE},
		)
		self.client.force_login(self.internal_user)

		response = self.client.get(reverse("puskesmas:stock_current"))

		self.assertEqual(response.status_code, 403)

	def test_search_and_mismatch_filters_do_not_widen_facility_scope(self):
		self.client.force_login(self.user)

		response = self.client.get(
			reverse("puskesmas:stock_current"),
			{
				"year": str(self.year),
				"q": "paracetamol",
				"mismatch": "mismatch",
				"facility": str(self.other_facility.pk),
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(len(response.context["rows"]), 1)
		self.assertEqual(response.context["rows"][0]["kode_barang"], "STK-001")
		self.assertEqual(response.context["stats"]["mismatch_count"], 1)
		self.assertNotContains(response, self.other_facility.name)

	def test_invalid_filters_render_errors_and_no_rows(self):
		self.client.force_login(self.user)

		response = self.client.get(
			reverse("puskesmas:stock_current"),
			{"year": "99999", "q": "bad\x00text", "mismatch": "bad"},
		)

		self.assertEqual(response.status_code, 200)
		self.assertTrue(response.context["form"].errors)
		self.assertEqual(response.context["rows"], [])
		self.assertEqual(response.context["stats"]["total_rows"], 0)

	def test_sidebar_link_visible_for_puskesmas_user(self):
		self.client.force_login(self.user)

		response = self.client.get(reverse("dashboard"))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, reverse("puskesmas:stock_current"))
		self.assertContains(response, "Stok Puskesmas")


class PuskesmasSBBKFormTests(TestCase):
	def setUp(self):
		self.unit = Unit.objects.create(code="TAB", name="Tablet")
		self.category = Category.objects.create(code="OBT", name="Obat", sort_order=1)
		self.facility = Facility.objects.create(
			code="PKM-01",
			name="Puskesmas Satu",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.other_facility = Facility.objects.create(
			code="PKM-02",
			name="Puskesmas Dua",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.item = Item.objects.create(
			nama_barang="Paracetamol",
			satuan=self.unit,
			kategori=self.category,
			is_active=True,
		)
		self.other_item = Item.objects.create(
			nama_barang="Amoxicillin",
			satuan=self.unit,
			kategori=self.category,
			is_active=True,
		)
		self.user = User.objects.create_user(
			username="operator-sbbk",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=self.facility,
		)
		self.distribution = Distribution.objects.create(
			distribution_type=Distribution.DistributionType.LPLPO,
			request_date="2026-04-01",
			distributed_date="2026-04-02",
			facility=self.facility,
			status=Distribution.Status.DISTRIBUTED,
			created_by=self.user,
		)
		self.distribution_item = DistributionItem.objects.create(
			distribution=self.distribution,
			item=self.item,
			quantity_requested=Decimal("5.00"),
			quantity_approved=Decimal("5.00"),
			issued_batch_lot="BATCH-01",
			issued_unit_price=Decimal("1000.00"),
		)

	def test_sbbk_number_is_auto_generated(self):
		sbbk = PuskesmasSBBK.objects.create(
			facility=self.facility,
			received_date="2026-04-02",
			created_by=self.user,
		)

		self.assertRegex(sbbk.document_number, r"^RCVCONF-\d{6}-\d{5}$")

	def test_sbbk_number_uses_received_date_month_prefix(self):
		sbbk = PuskesmasSBBK.objects.create(
			facility=self.facility,
			received_date="2024-02-15",
			created_by=self.user,
		)

		self.assertTrue(sbbk.document_number.startswith("RCVCONF-202402-"))

	def test_sbbk_number_generation_retries_after_unique_collision(self):
		sbbk = PuskesmasSBBK(
			facility=self.facility,
			received_date="2026-04-02",
			created_by=self.user,
		)

		with patch.object(
			PuskesmasSBBK,
			"_generate_next_document_number",
			side_effect=["RCVCONF-202604-00001", "RCVCONF-202604-00002"],
		), patch(
			"apps.puskesmas.models.TimeStampedModel.save",
			side_effect=[IntegrityError("duplicate key"), None],
		) as mocked_save:
			sbbk.save()

		self.assertEqual(sbbk.document_number, "RCVCONF-202604-00002")
		self.assertEqual(mocked_save.call_count, 2)

	def test_same_item_can_exist_multiple_times_in_one_sbbk(self):
		sbbk = PuskesmasSBBK.objects.create(
			facility=self.facility,
			received_date="2026-04-02",
			created_by=self.user,
		)
		PuskesmasSBBKItem.objects.create(
			sbbk=sbbk,
			item=self.item,
			quantity=Decimal("2.00"),
			unit_price=Decimal("1000.00"),
		)
		PuskesmasSBBKItem.objects.create(
			sbbk=sbbk,
			item=self.item,
			quantity=Decimal("3.00"),
			unit_price=Decimal("1200.00"),
		)

		self.assertEqual(sbbk.items.filter(item=self.item).count(), 2)

	def test_operator_sbbk_form_ignores_posted_other_facility(self):
		form = PuskesmasSBBKForm(
			data={
				"document_number": "  MANUAL-001  ",
				"facility": self.other_facility.pk,
				"distribution": self.distribution.pk,
				"received_date": "2026-04-02",
				"notes": "  Catatan  ",
			},
			user=self.user,
		)

		self.assertTrue(form.is_valid())
		self.assertEqual(form.cleaned_data["facility"], self.facility)
		self.assertEqual(form.cleaned_data["document_number"], "MANUAL-001")
		self.assertEqual(form.cleaned_data["notes"], "Catatan")

	def test_sbbk_form_rejects_null_byte_strings(self):
		form = PuskesmasSBBKForm(
			data={
				"document_number": "BAD\x00DOC",
				"facility": self.facility.pk,
				"distribution": self.distribution.pk,
				"received_date": "2026-04-02",
				"notes": "",
			},
			user=self.user,
		)

		self.assertFalse(form.is_valid())
		self.assertIn("document_number", form.errors)

	def test_legacy_edit_form_allows_missing_distribution(self):
		legacy_sbbk = PuskesmasSBBK.objects.create(
			facility=self.facility,
			received_date="2026-04-02",
			created_by=self.user,
		)
		form = PuskesmasSBBKForm(
			data={
				"document_number": legacy_sbbk.document_number,
				"facility": self.facility.pk,
				"distribution": "",
				"received_date": "2026-04-03",
				"notes": "  Koreksi dokumen lama  ",
			},
			instance=legacy_sbbk,
			user=self.user,
		)

		self.assertTrue(form.is_valid())
		self.assertIsNone(form.cleaned_data["distribution"])
		self.assertEqual(form.cleaned_data["notes"], "Koreksi dokumen lama")

	def test_legacy_edit_form_rejects_new_distribution_link(self):
		legacy_sbbk = PuskesmasSBBK.objects.create(
			facility=self.facility,
			received_date="2026-04-02",
			created_by=self.user,
		)
		form = PuskesmasSBBKForm(
			data={
				"document_number": legacy_sbbk.document_number,
				"facility": self.facility.pk,
				"distribution": self.distribution.pk,
				"received_date": "2026-04-03",
				"notes": "Koreksi dokumen lama",
			},
			instance=legacy_sbbk,
			user=self.user,
		)

		self.assertFalse(form.is_valid())
		self.assertIn("distribution", form.errors)

	def test_create_form_still_requires_distribution(self):
		form = PuskesmasSBBKForm(
			data={
				"document_number": "",
				"facility": self.facility.pk,
				"distribution": "",
				"received_date": "2026-04-02",
				"notes": "",
			},
			user=self.user,
		)

		self.assertFalse(form.is_valid())
		self.assertIn("distribution", form.errors)

	def test_sbbk_item_form_rejects_non_finite_quantity_and_unit_price(self):
		for quantity, unit_price in (("NaN", "1000"), ("1", "Infinity")):
			with self.subTest(quantity=quantity, unit_price=unit_price):
				form = PuskesmasSBBKItemForm(
					data={
						"distribution_item": self.distribution_item.pk,
						"item": self.item.pk,
						"quantity": quantity,
						"unit_price": unit_price,
						"batch_lot": "BATCH-01",
						"expiry_date": "",
						"notes": "",
					},
					distribution=self.distribution,
				)
				self.assertFalse(form.is_valid())

	def test_sbbk_item_form_rejects_fractional_quantity(self):
		form = PuskesmasSBBKItemForm(
			data={
				"distribution_item": self.distribution_item.pk,
				"item": self.item.pk,
				"quantity": "0.50",
				"unit_price": "1000",
				"batch_lot": "BATCH-01",
				"expiry_date": "",
				"notes": "",
			},
			distribution=self.distribution,
		)

		self.assertFalse(form.is_valid())
		self.assertIn("quantity", form.errors)

	def test_legacy_sbbk_item_form_allows_missing_distribution_item(self):
		legacy_sbbk = PuskesmasSBBK.objects.create(
			facility=self.facility,
			received_date="2026-04-02",
			created_by=self.user,
		)
		legacy_item = PuskesmasSBBKItem.objects.create(
			sbbk=legacy_sbbk,
			item=self.item,
			quantity=Decimal("2.00"),
			unit_price=Decimal("1000.00"),
			batch_lot="LEGACY-01",
		)
		form = PuskesmasSBBKItemForm(
			data={
				"distribution_item": "",
				"item": self.item.pk,
				"quantity": "3",
				"unit_price": "1100.00",
				"batch_lot": " LEGACY-02 ",
				"expiry_date": "",
				"notes": "  Pembetulan arsip  ",
			},
			instance=legacy_item,
			distribution=None,
		)

		self.assertTrue(form.is_valid())
		self.assertIsNone(form.cleaned_data["distribution_item"])
		self.assertEqual(form.cleaned_data["item"], self.item)
		self.assertEqual(form.cleaned_data["batch_lot"], "LEGACY-02")

	def test_linked_sbbk_item_form_still_requires_distribution_item(self):
		form = PuskesmasSBBKItemForm(
			data={
				"distribution_item": "",
				"item": self.item.pk,
				"quantity": "1",
				"unit_price": "1000",
				"batch_lot": "BATCH-01",
				"expiry_date": "",
				"notes": "",
			},
			distribution=self.distribution,
		)

		self.assertFalse(form.is_valid())
		self.assertIn("distribution_item", form.errors)

	def test_sbbk_item_model_rejects_fractional_quantity(self):
		sbbk = PuskesmasSBBK.objects.create(
			facility=self.facility,
			received_date="2026-04-02",
			created_by=self.user,
		)
		item = PuskesmasSBBKItem(
			sbbk=sbbk,
			item=self.item,
			quantity=Decimal("1.50"),
			unit_price=Decimal("1000.00"),
		)

		with self.assertRaises(ValidationError):
			item.full_clean()

	def test_sbbk_item_form_requires_adjustment_note_for_create_time_quantity_difference(self):
		form = PuskesmasSBBKItemForm(
			data={
				"distribution_item": self.distribution_item.pk,
				"item": self.item.pk,
				"quantity": "4",
				"unit_price": "1000.00",
				"batch_lot": "BATCH-01",
				"expiry_date": "",
				"notes": "",
			},
			distribution=self.distribution,
		)

		self.assertFalse(form.is_valid())
		self.assertIn("notes", form.errors)

	def test_sbbk_item_model_requires_adjustment_note_with_unsaved_parent_distribution(self):
		sbbk = PuskesmasSBBK(
			facility=self.facility,
			distribution=self.distribution,
			received_date="2026-04-02",
			created_by=self.user,
		)
		item = PuskesmasSBBKItem(
			sbbk=sbbk,
			distribution_item=self.distribution_item,
			item=self.item,
			quantity=Decimal("4.00"),
			unit_price=Decimal("1000.00"),
			batch_lot="OTHER-BATCH",
		)

		with self.assertRaises(ValidationError) as exc_info:
			item.full_clean()

		self.assertIn("notes", exc_info.exception.message_dict)


class PuskesmasReceiptConfirmationStatusMigrationTests(TransactionTestCase):
	migrate_from = ("puskesmas", "0006_alter_puskesmasreceiptconfirmation_created_by_and_more")
	migrate_to = ("puskesmas", "0008_backfill_receipt_confirmation_status")

	def setUp(self):
		super().setUp()
		self.executor = MigrationExecutor(connection)
		self.executor.migrate([self.migrate_from])

		old_apps = self.executor.loader.project_state([self.migrate_from]).apps
		ReceiptConfirmationModel = old_apps.get_model(
			"puskesmas",
			"PuskesmasReceiptConfirmation",
		)

		facility = Facility.objects.create(
			code="MIG-PKM",
			name="Puskesmas Migrasi",
			facility_type="PUSKESMAS",
			is_active=True,
		)
		user = User.objects.create_user(
			username="migration-operator",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=facility,
		)
		self.receipt_pk = ReceiptConfirmationModel.objects.create(
			document_number="RCVCONF-202602-00001",
			facility_id=facility.pk,
			received_date="2026-02-10",
			notes="Data lama sebelum field status ditambahkan.",
			created_by_id=user.pk,
		).pk

		self.executor = MigrationExecutor(connection)
		self.executor.migrate([self.migrate_to])
		self.apps = self.executor.loader.project_state([self.migrate_to]).apps

	def tearDown(self):
		executor = MigrationExecutor(connection)
		executor.migrate(executor.loader.graph.leaf_nodes())
		super().tearDown()

	def test_existing_receipts_are_backfilled_to_confirmed(self):
		ReceiptConfirmationModel = self.apps.get_model(
			"puskesmas",
			"PuskesmasReceiptConfirmation",
		)
		receipt = ReceiptConfirmationModel.objects.get(pk=self.receipt_pk)

		self.assertEqual(receipt.status, "CONFIRMED")


class PuskesmasConsumptionFormTests(TestCase):
	def setUp(self):
		self.unit = Unit.objects.create(code="TAB", name="Tablet")
		self.category = Category.objects.create(code="OBT", name="Obat", sort_order=1)
		self.facility = Facility.objects.create(
			code="PKM-CNS-01",
			name="Puskesmas Konsumsi",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.item = Item.objects.create(
			nama_barang="Paracetamol",
			satuan=self.unit,
			kategori=self.category,
			is_active=True,
		)
		self.subunit = PuskesmasSubunit.objects.create(
			facility=self.facility,
			name="Poli Umum",
			subunit_type=PuskesmasSubunit.SubunitType.TREATMENT_ROOM,
		)
		self.user = User.objects.create_user(
			username="operator-consumption-form",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=self.facility,
		)

	def test_subunit_form_ignores_posted_other_facility(self):
		other_facility = Facility.objects.create(
			code="PKM-CNS-02",
			name="Puskesmas Lain",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		form = PuskesmasSubunitForm(
			data={
				"facility": other_facility.pk,
				"name": "  Poli Gigi  ",
				"subunit_type": PuskesmasSubunit.SubunitType.TREATMENT_ROOM,
				"sort_order": "0",
				"is_active": "on",
			},
			user=self.user,
		)

		self.assertTrue(form.is_valid())
		self.assertEqual(form.cleaned_data["facility"], self.facility)
		self.assertEqual(form.cleaned_data["name"], "Poli Gigi")

	def test_subunit_form_uses_indonesian_labels_and_hover_titles(self):
		form = PuskesmasSubunitForm(user=self.user)

		self.assertEqual(form.fields["name"].label, "Nama Poli/Pustu")
		self.assertEqual(form.fields["subunit_type"].label, "Jenis Poli/Pustu")
		self.assertIn("poli atau pustu", form.fields["name"].widget.attrs["title"].lower())
		self.assertNotIn("sort_order", form.fields)

	def test_subunit_form_rejects_facility_change_after_consumption_history_exists(self):
		admin = User.objects.create_superuser(
			username="admin-subunit-form",
			email="admin-subunit-form@example.com",
			password="TestPassword123!",
		)
		other_facility = Facility.objects.create(
			code="PKM-CNS-03",
			name="Puskesmas Pindah",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		consumption = PuskesmasConsumption.objects.create(
			facility=self.facility,
			bulan=2,
			tahun=2026,
			created_by=self.user,
		)
		PuskesmasConsumptionEntry.objects.create(
			consumption=consumption,
			item=self.item,
			subunit=self.subunit,
			quantity=3,
		)

		form = PuskesmasSubunitForm(
			data={
				"facility": other_facility.pk,
				"name": self.subunit.name,
				"subunit_type": self.subunit.subunit_type,
				"sort_order": str(self.subunit.sort_order),
				"is_active": "on",
			},
			instance=self.subunit,
			user=admin,
		)

		self.assertFalse(form.is_valid())
		self.assertIn("facility", form.errors)

	def test_consumption_form_rejects_fractional_matrix_quantity(self):
		form = PuskesmasConsumptionMatrixForm(
			data={
				"facility": self.facility.pk,
				"bulan": "2",
				"tahun": "2026",
				"notes": "",
				f"qty_{self.item.pk}_{self.subunit.pk}": "1.50",
			},
			user=self.user,
			subunits=[self.subunit],
			items=[self.item],
		)

		self.assertFalse(form.is_valid())
		self.assertIn(f"qty_{self.item.pk}_{self.subunit.pk}", form.errors)

	def test_consumption_form_adds_hover_information_to_matrix_inputs(self):
		form = PuskesmasConsumptionMatrixForm(
			user=self.user,
			subunits=[self.subunit],
			items=[self.item],
		)

		field = form.fields[f"qty_{self.item.pk}_{self.subunit.pk}"]
		self.assertEqual(form.fields["bulan"].label, "Bulan")
		self.assertIn("periode pemakaian", form.fields["bulan"].widget.attrs["title"].lower())
		self.assertEqual(field.widget.attrs["placeholder"], "0")
		self.assertIn(self.subunit.name, field.widget.attrs["title"])

	def test_assert_consumption_month_mutable_uses_row_lock_when_requested(self):
		lplpo = LPLPO.objects.create(
			facility=self.facility,
			bulan=2,
			tahun=2026,
			status=LPLPO.Status.DRAFT,
			created_by=self.user,
		)

		with patch("apps.puskesmas.services.LPLPO.objects.filter") as mocked_filter:
			mocked_queryset = mocked_filter.return_value
			mocked_queryset.select_for_update.return_value = mocked_queryset
			mocked_queryset.only.return_value.first.return_value = lplpo

			result = assert_consumption_month_mutable(
				facility=self.facility,
				bulan=2,
				tahun=2026,
				lock=True,
			)

		mocked_filter.assert_called_once_with(
			facility=self.facility,
			bulan=2,
			tahun=2026,
		)
		mocked_queryset.select_for_update.assert_called_once_with()
		self.assertEqual(result, lplpo)

	def test_consumption_subunits_follow_creation_time_order(self):
		older_subunit = PuskesmasSubunit.objects.create(
			facility=self.facility,
			name="Poli Z",
			subunit_type=PuskesmasSubunit.SubunitType.TREATMENT_ROOM,
		)
		newer_subunit = PuskesmasSubunit.objects.create(
			facility=self.facility,
			name="Poli A",
			subunit_type=PuskesmasSubunit.SubunitType.TREATMENT_ROOM,
		)
		now = timezone.now()
		PuskesmasSubunit.objects.filter(pk=older_subunit.pk).update(
			created_at=now - timedelta(minutes=2)
		)
		PuskesmasSubunit.objects.filter(pk=newer_subunit.pk).update(
			created_at=now - timedelta(minutes=1)
		)

		subunits = _get_consumption_subunits(self.facility)

		self.assertEqual([subunit.pk for subunit in subunits[:2]], [older_subunit.pk, newer_subunit.pk])


class PuskesmasSBBKViewTests(SecureClientDefaultsMixin, TestCase):
	def setUp(self):
		super().setUp()
		self.unit = Unit.objects.create(code="TAB", name="Tablet")
		self.category = Category.objects.create(code="OBT", name="Obat", sort_order=1)
		self.facility = Facility.objects.create(
			code="PKM-01",
			name="Puskesmas Satu",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.other_facility = Facility.objects.create(
			code="PKM-02",
			name="Puskesmas Dua",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.item = Item.objects.create(
			nama_barang="Paracetamol",
			satuan=self.unit,
			kategori=self.category,
			is_active=True,
		)
		self.other_item = Item.objects.create(
			nama_barang="Amoxicillin",
			satuan=self.unit,
			kategori=self.category,
			is_active=True,
		)
		self.operator = User.objects.create_user(
			username="operator-sbbk-view",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=self.facility,
		)
		self.other_operator = User.objects.create_user(
			username="operator-sbbk-other",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=self.other_facility,
		)
		self.operator_without_facility = User.objects.create_user(
			username="operator-sbbk-no-facility",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
		)
		ensure_default_module_access(self.operator_without_facility, overwrite=True)
		self.admin = User.objects.create_superuser(
			username="admin-sbbk",
			email="admin-sbbk@example.com",
			password="TestPassword123!",
		)

	def _create_lplpo(self, *, facility=None, bulan=2, tahun=2026, status=LPLPO.Status.DRAFT, items=None):
		facility = facility or self.facility
		lplpo = LPLPO.objects.create(
			facility=facility,
			bulan=bulan,
			tahun=tahun,
			status=status,
			created_by=self.operator if facility == self.facility else self.other_operator,
		)
		for item in items or [self.item]:
			LPLPOItem.objects.create(
				lplpo=lplpo,
				item=item,
				stock_awal=Decimal("5.00"),
				penerimaan=Decimal("0.00"),
				harga_satuan=Decimal("900.00"),
			)
		return lplpo

	def _create_sbbk(self, *, facility=None, received_date="2026-02-10", created_by=None, status=PuskesmasSBBK.ReceiptStatus.CONFIRMED):
		facility = facility or self.facility
		created_by = created_by or (self.operator if facility == self.facility else self.other_operator)
		return PuskesmasSBBK.objects.create(
			facility=facility,
			received_date=received_date,
			status=status,
			created_by=created_by,
		)

	def _create_distribution(self, *, facility=None, distributed_date="2026-02-10", created_by=None):
		facility = facility or self.facility
		created_by = created_by or (self.operator if facility == self.facility else self.other_operator)
		distribution = Distribution.objects.create(
			distribution_type=Distribution.DistributionType.LPLPO,
			request_date=distributed_date,
			distributed_date=distributed_date,
			facility=facility,
			status=Distribution.Status.DISTRIBUTED,
			created_by=created_by,
		)
		distribution_item = DistributionItem.objects.create(
			distribution=distribution,
			item=self.item,
			quantity_requested=Decimal("4.00"),
			quantity_approved=Decimal("4.00"),
			issued_batch_lot="BATCH-01",
			issued_unit_price=Decimal("1000.00"),
		)
		return distribution, distribution_item

	def _add_distribution_item(
		self,
		distribution,
		*,
		item=None,
		quantity="2.00",
		batch_lot="BATCH-02",
		unit_price="1200.00",
		expiry_date=None,
	):
		return DistributionItem.objects.create(
			distribution=distribution,
			item=item or self.other_item,
			quantity_requested=Decimal(quantity),
			quantity_approved=Decimal(quantity),
			issued_batch_lot=batch_lot,
			issued_unit_price=Decimal(unit_price),
			issued_expiry_date=expiry_date,
		)

	def _create_payload(
		self,
		*,
		distribution,
		distribution_items,
		facility=None,
		received_date="2026-02-10",
		notes="",
		confirmed_item_ids=None,
		action="save",
	):
		facility = facility or self.facility
		distribution_items = list(distribution_items)
		confirmed_item_ids = set(
			confirmed_item_ids
			if confirmed_item_ids is not None
			else [item.pk for item in distribution_items]
		)
		payload = {
			"action": action,
			"document_number": "",
			"facility": str(facility.pk),
			"distribution": str(distribution.pk),
			"received_date": received_date,
			"notes": notes,
			"items-TOTAL_FORMS": str(len(distribution_items)),
			"items-INITIAL_FORMS": "0",
			"items-MIN_NUM_FORMS": "1",
			"items-MAX_NUM_FORMS": "1000",
		}
		for index, distribution_item in enumerate(distribution_items):
			payload[f"items-{index}-distribution_item"] = str(distribution_item.pk)
			if distribution_item.pk in confirmed_item_ids:
				payload[f"items-{index}-confirmed"] = "on"
		return payload

	def _linked_edit_payload(
		self,
		receipt_confirmation,
		*,
		distribution_items,
		facility=None,
		received_date="2026-02-10",
		notes="",
		confirmed_item_ids=None,
	):
		facility = facility or self.facility
		distribution = receipt_confirmation.distribution
		distribution_items = list(distribution_items)
		confirmed_item_ids = set(
			confirmed_item_ids
			if confirmed_item_ids is not None
			else [item.pk for item in distribution_items]
		)
		payload = {
			"action": "draft",
			"document_number": receipt_confirmation.document_number,
			"facility": str(facility.pk),
			"distribution": str(distribution.pk),
			"received_date": received_date,
			"notes": notes,
			"items-TOTAL_FORMS": str(len(distribution_items)),
			"items-INITIAL_FORMS": "0",
			"items-MIN_NUM_FORMS": "1",
			"items-MAX_NUM_FORMS": "1000",
		}
		for index, distribution_item in enumerate(distribution_items):
			payload[f"items-{index}-distribution_item"] = str(distribution_item.pk)
			if distribution_item.pk in confirmed_item_ids:
				payload[f"items-{index}-confirmed"] = "on"
		return payload

	def _edit_payload(self, sbbk_item, *, facility=None, received_date="2026-02-10", quantity="4.00", unit_price="1000.00", notes="", distribution=None, distribution_item=None):
		facility = facility or self.facility
		distribution = distribution if distribution is not None else sbbk_item.sbbk.distribution
		distribution_item = (
			distribution_item if distribution_item is not None else sbbk_item.distribution_item
		)
		return {
			"status": PuskesmasSBBK.ReceiptStatus.CONFIRMED,
			"document_number": sbbk_item.sbbk.document_number,
			"facility": str(facility.pk),
			"distribution": str(distribution.pk) if distribution is not None else "",
			"received_date": received_date,
			"notes": "",
			"items-TOTAL_FORMS": "1",
			"items-INITIAL_FORMS": "1",
			"items-MIN_NUM_FORMS": "1",
			"items-MAX_NUM_FORMS": "1000",
			"items-0-id": str(sbbk_item.pk),
			"items-0-distribution_item": (
				str(distribution_item.pk) if distribution_item is not None else ""
			),
			"items-0-item": str(self.item.pk),
			"items-0-quantity": quantity,
			"items-0-unit_price": unit_price,
			"items-0-batch_lot": sbbk_item.batch_lot or "BATCH-01",
			"items-0-expiry_date": "",
			"items-0-notes": notes,
		}

	def test_operator_can_create_and_list_own_facility_sbbk(self):
		self.client.force_login(self.operator)
		distribution, distribution_item = self._create_distribution()
		create_response = self.client.post(
			reverse("puskesmas:receiving_create"),
			self._create_payload(
				distribution=distribution,
				distribution_items=[distribution_item],
				action="confirm",
			),
		)

		self.assertEqual(create_response.status_code, 302)
		sbbk = PuskesmasSBBK.objects.get()
		self.assertEqual(sbbk.facility, self.facility)
		self.assertEqual(sbbk.created_by, self.operator)

		list_response = self.client.get(reverse("puskesmas:receiving_list"))
		self.assertEqual(list_response.status_code, 200)
		self.assertContains(list_response, sbbk.document_number)

	def test_non_superuser_without_facility_gets_403_on_receiving_list(self):
		self.client.force_login(self.operator_without_facility)

		response = self.client.get(reverse("puskesmas:receiving_list"))

		self.assertEqual(response.status_code, 403)

	def test_cross_facility_receiving_detail_gets_403(self):
		sbbk = self._create_sbbk(facility=self.other_facility)

		self.client.force_login(self.operator)
		response = self.client.get(reverse("puskesmas:receiving_detail", args=[sbbk.pk]))

		self.assertEqual(response.status_code, 403)

	def test_create_get_preloads_distribution_rows(self):
		self.client.force_login(self.operator)
		distribution, distribution_item = self._create_distribution()

		response = self.client.get(
			reverse("puskesmas:receiving_create"),
			{"distribution": str(distribution.pk)},
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, distribution.document_number)
		self.assertContains(
			response,
			f'name="items-0-distribution_item" value="{distribution_item.pk}"',
			html=False,
		)
		self.assertContains(response, "Checklist Konfirmasi Penerimaan")
		self.assertNotContains(response, "Informasi Konfirmasi")
		self.assertContains(response, 'id="distribution-preview-form"', html=False)
		self.assertNotContains(response, 'formmethod="get"', html=False)

	def test_create_get_ignores_cross_facility_distribution_preview(self):
		self.client.force_login(self.operator)
		distribution, distribution_item = self._create_distribution(
			facility=self.other_facility,
		)

		response = self.client.get(
			reverse("puskesmas:receiving_create"),
			{"distribution": str(distribution.pk)},
		)

		self.assertEqual(response.status_code, 200)
		self.assertNotContains(response, distribution.document_number)
		self.assertNotContains(
			response,
			f'name="items-0-distribution_item" value="{distribution_item.pk}"',
			html=False,
		)

	def test_create_is_blocked_when_same_month_lplpo_is_submitted(self):
		self._create_lplpo(status=LPLPO.Status.SUBMITTED)
		distribution, distribution_item = self._create_distribution()
		self.client.force_login(self.operator)

		response = self.client.post(
			reverse("puskesmas:receiving_create"),
			self._create_payload(
				distribution=distribution,
				distribution_items=[distribution_item],
				action="confirm",
			),
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Konfirmasi penerimaan untuk periode ini tidak dapat diubah")
		self.assertEqual(PuskesmasSBBK.objects.count(), 0)

	@override_settings(
		PUSKESMAS_RECEIPT_CONFIRMATION_MUTATION_RATE_LIMIT="1/m",
		RATELIMIT_USE_CACHE="locmem",
	)
	def test_create_returns_429_when_rate_limited(self):
		self.client.force_login(self.operator)
		distribution_one, distribution_item_one = self._create_distribution(distributed_date="2026-02-10")
		distribution_two, distribution_item_two = self._create_distribution(distributed_date="2026-02-11")
		first_response = self.client.post(
			reverse("puskesmas:receiving_create"),
			self._create_payload(
				distribution=distribution_one,
				distribution_items=[distribution_item_one],
				action="confirm",
			),
		)
		second_response = self.client.post(
			reverse("puskesmas:receiving_create"),
			self._create_payload(
				distribution=distribution_two,
				distribution_items=[distribution_item_two],
				action="confirm",
			),
		)

		self.assertEqual(first_response.status_code, 302)
		self.assertEqual(second_response.status_code, 429)
		self.assertContains(
			second_response,
			"Terlalu banyak percobaan pada aksi ini",
			status_code=429,
		)

	@override_settings(
		PUSKESMAS_RECEIPT_CONFIRMATION_MUTATION_RATE_LIMIT="1/m",
		RATELIMIT_USE_CACHE="locmem",
	)
	def test_create_get_preview_does_not_consume_mutation_rate_limit(self):
		self.client.force_login(self.operator)
		distribution, distribution_item = self._create_distribution()

		first_preview = self.client.get(
			reverse("puskesmas:receiving_create"),
			{"distribution": str(distribution.pk)},
		)
		second_preview = self.client.get(
			reverse("puskesmas:receiving_create"),
			{"distribution": str(distribution.pk)},
		)
		save_response = self.client.post(
			reverse("puskesmas:receiving_create"),
			self._create_payload(
				distribution=distribution,
				distribution_items=[distribution_item],
				action="confirm",
			),
		)

		self.assertEqual(first_preview.status_code, 200)
		self.assertEqual(second_preview.status_code, 200)
		self.assertEqual(save_response.status_code, 302)

	def test_create_confirmed_receipt_recomputes_same_month_draft_lplpo(self):
		lplpo = self._create_lplpo()
		distribution, distribution_item = self._create_distribution()
		self.client.force_login(self.operator)

		response = self.client.post(
			reverse("puskesmas:receiving_create"),
			self._create_payload(
				distribution=distribution,
				distribution_items=[distribution_item],
				action="confirm",
			),
		)

		self.assertEqual(response.status_code, 302)
		self.assertEqual(PuskesmasSBBK.objects.get().status, PuskesmasSBBK.ReceiptStatus.CONFIRMED)
		line = lplpo.items.get(item=self.item)
		self.assertEqual(line.penerimaan, Decimal("4.00"))
		self.assertEqual(line.harga_satuan, Decimal("1000.00"))
		self.assertTrue(line.penerimaan_auto_filled)

	def test_create_draft_allows_partial_unchecked_rows(self):
		distribution, first_distribution_item = self._create_distribution()
		second_distribution_item = self._add_distribution_item(distribution)
		self.client.force_login(self.operator)

		response = self.client.post(
			reverse("puskesmas:receiving_create"),
			self._create_payload(
				distribution=distribution,
				distribution_items=[first_distribution_item, second_distribution_item],
				confirmed_item_ids={first_distribution_item.pk},
			),
		)

		self.assertEqual(response.status_code, 302)
		sbbk = PuskesmasSBBK.objects.get()
		self.assertEqual(sbbk.status, PuskesmasSBBK.ReceiptStatus.DRAFT)
		self.assertEqual(sbbk.items.count(), 1)

	def test_create_draft_allows_submission_when_no_rows_are_checked(self):
		distribution, first_distribution_item = self._create_distribution()
		second_distribution_item = self._add_distribution_item(distribution)
		self.client.force_login(self.operator)

		response = self.client.post(
			reverse("puskesmas:receiving_create"),
			self._create_payload(
				distribution=distribution,
				distribution_items=[first_distribution_item, second_distribution_item],
				confirmed_item_ids=set(),
			),
		)

		self.assertEqual(response.status_code, 302)
		sbbk = PuskesmasSBBK.objects.get()
		self.assertEqual(sbbk.status, PuskesmasSBBK.ReceiptStatus.DRAFT)
		self.assertEqual(sbbk.items.count(), 0)

	def test_create_partial_checklist_saves_draft_without_lplpo_sync(self):
		lplpo = self._create_lplpo(items=[self.item, self.other_item])
		distribution, first_distribution_item = self._create_distribution()
		second_distribution_item = self._add_distribution_item(
			distribution,
			item=self.other_item,
			quantity="3.00",
			batch_lot="BATCH-02",
			unit_price="1500.00",
		)
		self.client.force_login(self.operator)

		response = self.client.post(
			reverse("puskesmas:receiving_create"),
			self._create_payload(
				distribution=distribution,
				distribution_items=[first_distribution_item, second_distribution_item],
				confirmed_item_ids={first_distribution_item.pk},
				notes="Sebagian barang belum diterima dan akan ditindaklanjuti di luar sistem.",
			),
		)

		self.assertEqual(response.status_code, 302)
		sbbk = PuskesmasSBBK.objects.get()
		self.assertEqual(sbbk.status, PuskesmasSBBK.ReceiptStatus.DRAFT)
		self.assertEqual(sbbk.items.count(), 1)
		self.assertEqual(sbbk.items.get().distribution_item, first_distribution_item)
		self.assertEqual(
			lplpo.items.get(item=self.item).penerimaan,
			Decimal("0.00"),
		)
		self.assertEqual(
			lplpo.items.get(item=self.other_item).penerimaan,
			Decimal("0.00"),
		)

	def test_confirm_button_rejects_partial_checklist(self):
		distribution, first_distribution_item = self._create_distribution()
		second_distribution_item = self._add_distribution_item(distribution)
		self.client.force_login(self.operator)

		response = self.client.post(
			reverse("puskesmas:receiving_create"),
			self._create_payload(
				distribution=distribution,
				distribution_items=[first_distribution_item, second_distribution_item],
				confirmed_item_ids={first_distribution_item.pk},
				action="confirm",
			),
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(
			response,
			"Semua baris harus dicentang untuk menyimpan konfirmasi final. Gunakan Simpan Draft bila barang belum lengkap.",
		)

	def test_edit_draft_recomputes_lplpo_back_to_zero(self):
		lplpo = self._create_lplpo(items=[self.item, self.other_item])
		distribution, first_distribution_item = self._create_distribution()
		second_distribution_item = self._add_distribution_item(
			distribution,
			item=self.other_item,
			quantity="3.00",
			batch_lot="BATCH-02",
			unit_price="1500.00",
		)
		sbbk = self._create_sbbk(status=PuskesmasSBBK.ReceiptStatus.CONFIRMED)
		sbbk.distribution = distribution
		sbbk.notes = "Semua barang awalnya dicentang."
		sbbk.save(update_fields=["distribution", "notes", "updated_at"])
		PuskesmasSBBKItem.objects.create(
			sbbk=sbbk,
			distribution_item=first_distribution_item,
			item=self.item,
			quantity=Decimal("4.00"),
			unit_price=Decimal("1000.00"),
			batch_lot="BATCH-01",
		)
		PuskesmasSBBKItem.objects.create(
			sbbk=sbbk,
			distribution_item=second_distribution_item,
			item=self.other_item,
			quantity=Decimal("3.00"),
			unit_price=Decimal("1500.00"),
			batch_lot="BATCH-02",
		)
		self.client.force_login(self.operator)

		response = self.client.post(
			reverse("puskesmas:receiving_edit", args=[sbbk.pk]),
			self._linked_edit_payload(
				sbbk,
				distribution_items=[first_distribution_item, second_distribution_item],
				confirmed_item_ids={first_distribution_item.pk},
				notes="Baris kedua belum diterima sesuai fisik.",
			),
		)

		self.assertEqual(response.status_code, 302)
		sbbk.refresh_from_db()
		self.assertEqual(sbbk.status, PuskesmasSBBK.ReceiptStatus.DRAFT)
		self.assertEqual(sbbk.items.count(), 1)
		self.assertEqual(
			lplpo.items.get(item=self.item).penerimaan,
			Decimal("0.00"),
		)
		self.assertEqual(
			lplpo.items.get(item=self.other_item).penerimaan,
			Decimal("0.00"),
		)

	def test_edit_confirmed_receipt_keeps_lplpo_synced(self):
		lplpo = self._create_lplpo(items=[self.item, self.other_item])
		distribution, first_distribution_item = self._create_distribution()
		second_distribution_item = self._add_distribution_item(
			distribution,
			item=self.other_item,
			quantity="3.00",
			batch_lot="BATCH-02",
			unit_price="1500.00",
		)
		sbbk = self._create_sbbk(status=PuskesmasSBBK.ReceiptStatus.CONFIRMED)
		sbbk.distribution = distribution
		sbbk.save(update_fields=["distribution", "status", "updated_at"])
		self.client.force_login(self.operator)

		payload = self._linked_edit_payload(
			sbbk,
			distribution_items=[first_distribution_item, second_distribution_item],
			confirmed_item_ids={first_distribution_item.pk, second_distribution_item.pk},
			notes="",
		)
		payload["action"] = "confirm"
		response = self.client.post(reverse("puskesmas:receiving_edit", args=[sbbk.pk]), payload)

		self.assertEqual(response.status_code, 302)
		sbbk.refresh_from_db()
		self.assertEqual(sbbk.status, PuskesmasSBBK.ReceiptStatus.CONFIRMED)
		self.assertEqual(lplpo.items.get(item=self.item).penerimaan, Decimal("4.00"))
		self.assertEqual(lplpo.items.get(item=self.other_item).penerimaan, Decimal("3.00"))

	def test_legacy_edit_page_loads_without_distribution_link(self):
		sbbk = self._create_sbbk()
		PuskesmasSBBKItem.objects.create(
			sbbk=sbbk,
			item=self.item,
			quantity=Decimal("2.00"),
			unit_price=Decimal("1000.00"),
			batch_lot="LEGACY-01",
		)
		self.client.force_login(self.operator)

		response = self.client.get(reverse("puskesmas:receiving_edit", args=[sbbk.pk]))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Dokumen lama tanpa distribusi sumber")

	def test_legacy_edit_recomputes_same_month_draft_lplpo(self):
		lplpo = self._create_lplpo()
		sbbk = self._create_sbbk()
		sbbk_item = PuskesmasSBBKItem.objects.create(
			sbbk=sbbk,
			item=self.item,
			quantity=Decimal("2.00"),
			unit_price=Decimal("1000.00"),
			batch_lot="LEGACY-01",
		)
		self.client.force_login(self.operator)

		response = self.client.post(
			reverse("puskesmas:receiving_edit", args=[sbbk.pk]),
			self._edit_payload(
				sbbk_item,
				quantity="5.00",
				unit_price="1500.00",
				notes="Koreksi legacy",
				distribution=None,
				distribution_item=None,
			),
		)

		self.assertEqual(response.status_code, 302)
		sbbk.refresh_from_db()
		sbbk_item.refresh_from_db()
		self.assertIsNone(sbbk.distribution)
		self.assertIsNone(sbbk_item.distribution_item)
		self.assertEqual(sbbk_item.quantity, Decimal("5.00"))
		line = lplpo.items.get(item=self.item)
		self.assertEqual(line.penerimaan, Decimal("5.00"))
		self.assertEqual(line.harga_satuan, Decimal("1500.00"))

	def test_legacy_edit_same_month_facility_change_resyncs_both_lplpo_documents(self):
		source_lplpo = self._create_lplpo()
		target_lplpo = self._create_lplpo(facility=self.other_facility)
		sbbk = self._create_sbbk()
		sbbk_item = PuskesmasSBBKItem.objects.create(
			sbbk=sbbk,
			item=self.item,
			quantity=Decimal("2.00"),
			unit_price=Decimal("1000.00"),
			batch_lot="LEGACY-01",
		)
		self.client.force_login(self.admin)

		response = self.client.post(
			reverse("puskesmas:receiving_edit", args=[sbbk.pk]),
			self._edit_payload(
				sbbk_item,
				facility=self.other_facility,
				quantity="5.00",
				unit_price="1500.00",
				notes="Pindah fasilitas legacy",
				distribution=None,
				distribution_item=None,
			),
		)

		self.assertEqual(response.status_code, 302)
		sbbk.refresh_from_db()
		self.assertEqual(sbbk.facility, self.other_facility)
		self.assertEqual(source_lplpo.items.get(item=self.item).penerimaan, Decimal("0.00"))
		target_line = target_lplpo.items.get(item=self.item)
		self.assertEqual(target_line.penerimaan, Decimal("5.00"))
		self.assertEqual(target_line.harga_satuan, Decimal("1500.00"))

	def test_legacy_edit_is_blocked_when_same_month_lplpo_is_submitted(self):
		self._create_lplpo(status=LPLPO.Status.SUBMITTED)
		sbbk = self._create_sbbk()
		sbbk_item = PuskesmasSBBKItem.objects.create(
			sbbk=sbbk,
			item=self.item,
			quantity=Decimal("2.00"),
			unit_price=Decimal("1000.00"),
			batch_lot="LEGACY-01",
		)
		self.client.force_login(self.operator)

		response = self.client.post(
			reverse("puskesmas:receiving_edit", args=[sbbk.pk]),
			self._edit_payload(
				sbbk_item,
				quantity="3.00",
				unit_price="1200.00",
				distribution=None,
				distribution_item=None,
			),
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Konfirmasi penerimaan untuk periode ini tidak dapat diubah")

	def test_legacy_edit_rejects_forged_distribution_link_submission(self):
		distribution, _distribution_item = self._create_distribution()
		sbbk = self._create_sbbk()
		sbbk_item = PuskesmasSBBKItem.objects.create(
			sbbk=sbbk,
			item=self.item,
			quantity=Decimal("2.00"),
			unit_price=Decimal("1000.00"),
			batch_lot="LEGACY-01",
		)
		self.client.force_login(self.operator)

		response = self.client.post(
			reverse("puskesmas:receiving_edit", args=[sbbk.pk]),
			self._edit_payload(
				sbbk_item,
				distribution=distribution,
				distribution_item=None,
			),
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(
			response,
			"Dokumen lama tidak boleh ditautkan ulang ke distribusi sumber baru.",
		)
		sbbk.refresh_from_db()
		self.assertIsNone(sbbk.distribution)

	def test_cross_facility_legacy_edit_gets_403(self):
		sbbk = self._create_sbbk(facility=self.other_facility)
		PuskesmasSBBKItem.objects.create(
			sbbk=sbbk,
			item=self.item,
			quantity=Decimal("2.00"),
			unit_price=Decimal("1000.00"),
		)
		self.client.force_login(self.operator)

		response = self.client.get(reverse("puskesmas:receiving_edit", args=[sbbk.pk]))

		self.assertEqual(response.status_code, 403)

	def test_delete_recomputes_same_month_draft_lplpo(self):
		lplpo = self._create_lplpo()
		distribution_keep, distribution_item_keep = self._create_distribution(distributed_date="2026-02-10")
		sbbk_keep = self._create_sbbk(received_date="2026-02-10")
		sbbk_keep.distribution = distribution_keep
		sbbk_keep.save(update_fields=["distribution", "notes", "updated_at"])
		PuskesmasSBBKItem.objects.create(
			sbbk=sbbk_keep,
			distribution_item=distribution_item_keep,
			item=self.item,
			quantity=Decimal("6.00"),
			unit_price=Decimal("1200.00"),
			batch_lot="BATCH-01",
			notes="Selisih batch pertama",
		)
		distribution_delete, distribution_item_delete = self._create_distribution(distributed_date="2026-02-11")
		sbbk_delete = self._create_sbbk(received_date="2026-02-11")
		sbbk_delete.distribution = distribution_delete
		sbbk_delete.save(update_fields=["distribution", "notes", "updated_at"])
		PuskesmasSBBKItem.objects.create(
			sbbk=sbbk_delete,
			distribution_item=distribution_item_delete,
			item=self.item,
			quantity=Decimal("4.00"),
			unit_price=Decimal("900.00"),
			batch_lot="BATCH-02",
			notes="Selisih batch kedua",
		)
		self.client.force_login(self.operator)

		response = self.client.post(
			reverse("puskesmas:receiving_delete", args=[sbbk_delete.pk])
		)

		self.assertEqual(response.status_code, 302)
		line = lplpo.items.get(item=self.item)
		self.assertEqual(line.penerimaan, Decimal("6.00"))
		self.assertEqual(line.harga_satuan, Decimal("1200.00"))

	def test_legacy_sbbk_view_permission_still_grants_receiving_access(self):
		viewer = User.objects.create_user(
			username="legacy-sbbk-viewer",
			password="TestPassword123!",
			role=User.Role.AUDITOR,
			facility=self.facility,
		)
		content_type = ContentType.objects.get(
			app_label="puskesmas",
			model="puskesmasreceiptconfirmation",
		)
		legacy_permission, _created = Permission.objects.get_or_create(
			content_type=content_type,
			codename="view_puskesmassbbk",
			defaults={"name": "Can view legacy puskesmas sbbk"},
		)
		legacy_permission.user_set.add(viewer)
		ModuleAccess.objects.update_or_create(
			user=viewer,
			module=ModuleAccess.Module.PUSKESMAS,
			defaults={"scope": ModuleAccess.Scope.NONE},
		)
		self.client.force_login(viewer)

		response = self.client.get(reverse("puskesmas:receiving_list"))

		self.assertEqual(response.status_code, 200)

	def test_receiving_detail_uses_receiving_back_link(self):
		distribution, _distribution_item = self._create_distribution()
		sbbk = self._create_sbbk()
		sbbk.distribution = distribution
		sbbk.save(update_fields=["distribution", "notes", "updated_at"])
		self.client.force_login(self.operator)

		response = self.client.get(reverse("puskesmas:receiving_detail", args=[sbbk.pk]))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, reverse("puskesmas:receiving_list"))
		self.assertContains(response, "Konfirmasi Penerimaan")


class PuskesmasRequestCreateViewTests(SecureClientDefaultsMixin, TestCase):
	def setUp(self):
		super().setUp()
		self.unit = Unit.objects.create(code="TAB", name="Tablet")
		self.category = Category.objects.create(code="OBT", name="Obat", sort_order=1)
		self.facility = Facility.objects.create(
			code="PKM-01",
			name="Puskesmas Satu",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.other_facility = Facility.objects.create(
			code="PKM-02",
			name="Puskesmas Dua",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.item = Item.objects.create(
			nama_barang="Haloperidol 5 mg",
			satuan=self.unit,
			kategori=self.category,
			is_active=True,
		)
		self.user = User.objects.create_user(
			username="operator-view",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=self.facility,
		)
		self.staff_user = User.objects.create_superuser(
			username="instalasi-admin",
			email="instalasi-admin@example.com",
			password="TestPassword123!",
		)
		self.other_user = User.objects.create_user(
			username="operator-other",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=self.other_facility,
		)
		self.staff_facility_user = User.objects.create_user(
			username="admin-umum-facility",
			password="TestPassword123!",
			role=User.Role.ADMIN_UMUM,
			facility=self.facility,
		)
		ensure_default_module_access(self.staff_facility_user, overwrite=True)

	def test_create_binds_request_to_logged_in_facility(self):
		self.client.force_login(self.user)

		response = self.client.post(
			reverse("puskesmas:request_create"),
			{
				"document_number": "",
				"facility": self.other_facility.pk,
				"request_date": "2026-04-02",
				"program": "",
				"notes": "Permintaan uji",
				"items-TOTAL_FORMS": "3",
				"items-INITIAL_FORMS": "0",
				"items-MIN_NUM_FORMS": "1",
				"items-MAX_NUM_FORMS": "1000",
				"items-0-item": str(self.item.pk),
				"items-0-quantity_requested": "5",
				"items-0-notes": "",
				"items-1-item": "",
				"items-1-quantity_requested": "",
				"items-1-notes": "",
				"items-2-item": "",
				"items-2-quantity_requested": "",
				"items-2-notes": "",
			},
		)

		self.assertEqual(response.status_code, 302)
		req = PuskesmasRequest.objects.get()
		self.assertEqual(req.facility, self.facility)
		self.assertEqual(req.created_by, self.user)

	def test_instalasi_farmasi_cannot_create_request(self):
		self.client.force_login(self.staff_user)

		response = self.client.get(reverse("puskesmas:request_create"))

		self.assertEqual(response.status_code, 403)

	def test_instalasi_farmasi_list_hides_create_button(self):
		self.client.force_login(self.staff_user)

		response = self.client.get(reverse("puskesmas:request_list"))

		self.assertEqual(response.status_code, 200)
		self.assertNotContains(response, 'Buat Permintaan')

	def test_edit_shows_operator_facility_as_readonly(self):
		req = PuskesmasRequest.objects.create(
			facility=self.facility,
			created_by=self.user,
		)

		self.client.force_login(self.user)
		response = self.client.get(reverse("puskesmas:request_edit", args=[req.pk]))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, self.facility.name)
		self.assertContains(response, 'type="hidden"', html=False)

	def test_edit_keeps_operator_facility_even_if_post_tampered(self):
		req = PuskesmasRequest.objects.create(
			facility=self.facility,
			created_by=self.user,
		)
		item_line = PuskesmasRequestItem.objects.create(
			request=req,
			item=self.item,
			quantity_requested=5,
		)

		self.client.force_login(self.user)
		response = self.client.post(
			reverse("puskesmas:request_edit", args=[req.pk]),
			{
				"document_number": req.document_number,
				"facility": self.other_facility.pk,
				"request_date": "2026-04-02",
				"program": "",
				"notes": "Permintaan edit",
				"items-TOTAL_FORMS": "1",
				"items-INITIAL_FORMS": "1",
				"items-MIN_NUM_FORMS": "1",
				"items-MAX_NUM_FORMS": "1000",
				"items-0-id": str(item_line.pk),
				"items-0-item": str(self.item.pk),
				"items-0-quantity_requested": "5",
				"items-0-notes": "",
			},
		)

		self.assertEqual(response.status_code, 302)
		req.refresh_from_db()
		self.assertEqual(req.facility, self.facility)

	def test_list_only_shows_operator_facility_requests(self):
		own_request = PuskesmasRequest.objects.create(
			facility=self.facility,
			created_by=self.user,
		)
		other_request = PuskesmasRequest.objects.create(
			facility=self.other_facility,
			created_by=self.other_user,
		)

		self.client.force_login(self.user)
		response = self.client.get(reverse("puskesmas:request_list"))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, own_request.document_number)
		self.assertNotContains(response, other_request.document_number)

	def test_detail_returns_403_when_operator_accesses_other_facility_request(self):
		other_request = PuskesmasRequest.objects.create(
			facility=self.other_facility,
			created_by=self.other_user,
		)

		self.client.force_login(self.user)
		response = self.client.get(
			reverse("puskesmas:request_detail", args=[other_request.pk])
		)

		self.assertEqual(response.status_code, 403)

	def test_edit_returns_403_when_operator_accesses_other_facility_request(self):
		other_request = PuskesmasRequest.objects.create(
			facility=self.other_facility,
			created_by=self.other_user,
		)

		self.client.force_login(self.user)
		response = self.client.get(
			reverse("puskesmas:request_edit", args=[other_request.pk])
		)

		self.assertEqual(response.status_code, 403)

	def test_submit_returns_403_when_operator_accesses_other_facility_request(self):
		other_request = PuskesmasRequest.objects.create(
			facility=self.other_facility,
			status=PuskesmasRequest.Status.DRAFT,
			created_by=self.other_user,
		)
		PuskesmasRequestItem.objects.create(
			request=other_request,
			item=self.item,
			quantity_requested=Decimal("2.00"),
		)

		self.client.force_login(self.user)
		response = self.client.post(
			reverse("puskesmas:request_submit", args=[other_request.pk])
		)

		self.assertEqual(response.status_code, 403)
		other_request.refresh_from_db()
		self.assertEqual(other_request.status, PuskesmasRequest.Status.DRAFT)

	def test_non_superuser_without_facility_cannot_view_request_list(self):
		kepala = User.objects.create_user(
			username="kepala-legacy",
			password="TestPassword123!",
			role=User.Role.KEPALA,
		)
		ModuleAccess.objects.filter(user=kepala).delete()
		PuskesmasRequest.objects.create(
			facility=self.facility,
			created_by=self.user,
		)

		self.client.force_login(kepala)
		response = self.client.get(reverse("puskesmas:request_list"))

		self.assertEqual(response.status_code, 403)

	def test_non_superuser_without_facility_cannot_view_request_detail(self):
		admin_umum = User.objects.create_user(
			username="admin-umum-legacy",
			password="TestPassword123!",
			role=User.Role.ADMIN_UMUM,
		)
		ModuleAccess.objects.filter(user=admin_umum).delete()
		req = PuskesmasRequest.objects.create(
			facility=self.facility,
			created_by=self.user,
		)

		self.client.force_login(admin_umum)
		response = self.client.get(reverse("puskesmas:request_detail", args=[req.pk]))

		self.assertEqual(response.status_code, 403)

	def test_non_superuser_with_facility_list_is_scoped(self):
		own_request = PuskesmasRequest.objects.create(
			facility=self.facility,
			created_by=self.user,
		)
		other_request = PuskesmasRequest.objects.create(
			facility=self.other_facility,
			created_by=self.other_user,
		)

		self.client.force_login(self.staff_facility_user)
		response = self.client.get(reverse("puskesmas:request_list"))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, own_request.document_number)
		self.assertNotContains(response, other_request.document_number)

	def test_non_superuser_with_facility_cannot_view_other_facility_request(self):
		other_request = PuskesmasRequest.objects.create(
			facility=self.other_facility,
			created_by=self.other_user,
		)

		self.client.force_login(self.staff_facility_user)
		response = self.client.get(reverse("puskesmas:request_detail", args=[other_request.pk]))

		self.assertEqual(response.status_code, 403)

	def test_explicit_none_scope_cannot_view_request_list(self):
		auditor = User.objects.create_user(
			username="auditor-legacy",
			password="TestPassword123!",
			role=User.Role.AUDITOR,
		)
		ModuleAccess.objects.update_or_create(
			user=auditor,
			module=ModuleAccess.Module.PUSKESMAS,
			defaults={"scope": ModuleAccess.Scope.NONE},
		)

		self.client.force_login(auditor)
		response = self.client.get(reverse("puskesmas:request_list"))

		self.assertEqual(response.status_code, 403)


class PuskesmasRequestApprovalTests(SecureClientDefaultsMixin, TestCase):
	def setUp(self):
		super().setUp()
		self.unit = Unit.objects.create(code="TAB", name="Tablet")
		self.category = Category.objects.create(code="OBT", name="Obat", sort_order=1)
		self.facility = Facility.objects.create(
			code="PKM-01",
			name="Puskesmas Satu",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.item = Item.objects.create(
			nama_barang="Haloperidol 5 mg",
			satuan=self.unit,
			kategori=self.category,
			is_active=True,
		)
		self.requester = User.objects.create_user(
			username="operator-submit",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=self.facility,
		)
		self.approver = User.objects.create_user(
			username="approver-root",
			password="TestPassword123!",
			role=User.Role.KEPALA,
			facility=self.facility,
		)
		self.manager = User.objects.create_superuser(
			username="manager-puskesmas",
			email="manager-puskesmas@example.com",
			password="TestPassword123!",
		)
		ModuleAccess.objects.update_or_create(
			user=self.approver,
			module=ModuleAccess.Module.PUSKESMAS,
			defaults={"scope": ModuleAccess.Scope.APPROVE},
		)
		ModuleAccess.objects.update_or_create(
			user=self.manager,
			module=ModuleAccess.Module.PUSKESMAS,
			defaults={"scope": ModuleAccess.Scope.MANAGE},
		)
		self.other_facility = Facility.objects.create(
			code="PKM-02",
			name="Puskesmas Dua",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.other_requester = User.objects.create_user(
			username="operator-submit-other",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=self.other_facility,
		)

	def test_approve_creates_distribution_with_requested_and_approved_quantities(self):
		req = PuskesmasRequest.objects.create(
			facility=self.facility,
			request_date="2026-04-02",
			status=PuskesmasRequest.Status.SUBMITTED,
			created_by=self.requester,
		)
		item_line = PuskesmasRequestItem.objects.create(
			request=req,
			item=self.item,
			quantity_requested=Decimal("8.00"),
		)

		self.client.force_login(self.approver)
		response = self.client.post(
			reverse("puskesmas:request_approve", args=[req.pk]),
			{
				f"approve_{item_line.pk}-quantity_approved": "5.00",
			},
		)

		self.assertEqual(response.status_code, 302)
		req.refresh_from_db()
		distribution = req.distribution
		self.assertIsNotNone(distribution)
		self.assertEqual(
			distribution.distribution_type,
			Distribution.DistributionType.SPECIAL_REQUEST,
		)
		self.assertTrue(distribution.staff_assignments.filter(user=self.approver).exists())

		line = distribution.items.get()
		self.assertEqual(line.quantity_requested, Decimal("8.00"))
		self.assertEqual(line.quantity_approved, Decimal("5.00"))

	def test_operator_detail_does_not_show_approval_actions(self):
		req = PuskesmasRequest.objects.create(
			facility=self.facility,
			request_date="2026-04-02",
			status=PuskesmasRequest.Status.SUBMITTED,
			created_by=self.requester,
		)
		PuskesmasRequestItem.objects.create(
			request=req,
			item=self.item,
			quantity_requested=Decimal("8.00"),
		)

		self.client.force_login(self.requester)
		response = self.client.get(reverse("puskesmas:request_detail", args=[req.pk]))

		self.assertEqual(response.status_code, 200)
		self.assertNotContains(response, "Setujui Permintaan")
		self.assertNotContains(response, "Tolak")

	def test_operator_cannot_approve_request_directly(self):
		req = PuskesmasRequest.objects.create(
			facility=self.facility,
			request_date="2026-04-02",
			status=PuskesmasRequest.Status.SUBMITTED,
			created_by=self.requester,
		)
		item_line = PuskesmasRequestItem.objects.create(
			request=req,
			item=self.item,
			quantity_requested=Decimal("8.00"),
		)

		self.client.force_login(self.requester)
		response = self.client.post(
			reverse("puskesmas:request_approve", args=[req.pk]),
			{
				f"approve_{item_line.pk}-quantity_approved": "5.00",
			},
		)

		self.assertEqual(response.status_code, 403)
		req.refresh_from_db()
		self.assertEqual(req.status, PuskesmasRequest.Status.SUBMITTED)
		self.assertIsNone(req.distribution)

	def test_operator_cannot_reject_request_directly(self):
		req = PuskesmasRequest.objects.create(
			facility=self.facility,
			request_date="2026-04-02",
			status=PuskesmasRequest.Status.SUBMITTED,
			created_by=self.requester,
		)

		self.client.force_login(self.requester)
		response = self.client.post(
			reverse("puskesmas:request_reject", args=[req.pk]),
			{"rejection_reason": "Tidak valid"},
		)

		self.assertEqual(response.status_code, 403)
		req.refresh_from_db()
		self.assertEqual(req.status, PuskesmasRequest.Status.SUBMITTED)

	def test_approver_can_reset_submitted_request_to_draft(self):
		req = PuskesmasRequest.objects.create(
			facility=self.facility,
			request_date="2026-04-02",
			status=PuskesmasRequest.Status.SUBMITTED,
			created_by=self.requester,
		)
		PuskesmasRequestItem.objects.create(
			request=req,
			item=self.item,
			quantity_requested=Decimal("4.00"),
		)

		self.client.force_login(self.approver)
		response = self.client.post(
			reverse("puskesmas:request_reset_draft", args=[req.pk])
		)

		self.assertEqual(response.status_code, 302)
		req.refresh_from_db()
		self.assertEqual(req.status, PuskesmasRequest.Status.DRAFT)

	def test_approver_detail_hides_draft_actions(self):
		req = PuskesmasRequest.objects.create(
			facility=self.facility,
			request_date="2026-04-02",
			status=PuskesmasRequest.Status.DRAFT,
			created_by=self.requester,
		)
		PuskesmasRequestItem.objects.create(
			request=req,
			item=self.item,
			quantity_requested=Decimal("4.00"),
		)

		self.client.force_login(self.approver)
		response = self.client.get(reverse("puskesmas:request_detail", args=[req.pk]))

		self.assertEqual(response.status_code, 200)
		self.assertNotContains(response, "Ajukan Permintaan")
		self.assertNotContains(response, 'id="edit-btn"', html=False)
		self.assertNotContains(response, 'id="delete-open-btn"', html=False)

	def test_approver_cannot_submit_draft_request(self):
		req = PuskesmasRequest.objects.create(
			facility=self.facility,
			request_date="2026-04-02",
			status=PuskesmasRequest.Status.DRAFT,
			created_by=self.requester,
		)
		PuskesmasRequestItem.objects.create(
			request=req,
			item=self.item,
			quantity_requested=Decimal("4.00"),
		)

		self.client.force_login(self.approver)
		response = self.client.post(reverse("puskesmas:request_submit", args=[req.pk]))

		self.assertEqual(response.status_code, 302)
		self.assertRedirects(response, reverse("puskesmas:request_detail", args=[req.pk]))
		req.refresh_from_db()
		self.assertEqual(req.status, PuskesmasRequest.Status.DRAFT)

	def test_manager_can_submit_draft_request(self):
		req = PuskesmasRequest.objects.create(
			facility=self.facility,
			request_date="2026-04-02",
			status=PuskesmasRequest.Status.DRAFT,
			created_by=self.requester,
		)
		PuskesmasRequestItem.objects.create(
			request=req,
			item=self.item,
			quantity_requested=Decimal("4.00"),
		)

		self.client.force_login(self.manager)
		response = self.client.post(reverse("puskesmas:request_submit", args=[req.pk]))

		self.assertEqual(response.status_code, 302)
		req.refresh_from_db()
		self.assertEqual(req.status, PuskesmasRequest.Status.SUBMITTED)

	def test_approver_cannot_approve_other_facility_request(self):
		req = PuskesmasRequest.objects.create(
			facility=self.other_facility,
			request_date="2026-04-02",
			status=PuskesmasRequest.Status.SUBMITTED,
			created_by=self.other_requester,
		)
		item_line = PuskesmasRequestItem.objects.create(
			request=req,
			item=self.item,
			quantity_requested=Decimal("8.00"),
		)

		self.client.force_login(self.approver)
		response = self.client.post(
			reverse("puskesmas:request_approve", args=[req.pk]),
			{f"approve_{item_line.pk}-quantity_approved": "5.00"},
		)

		self.assertEqual(response.status_code, 403)
		req.refresh_from_db()
		self.assertEqual(req.status, PuskesmasRequest.Status.SUBMITTED)
		self.assertIsNone(req.distribution)

class PuskesmasConsumptionViewTests(SecureClientDefaultsMixin, TestCase):
	def setUp(self):
		super().setUp()
		self.unit = Unit.objects.create(code="TAB", name="Tablet")
		self.category = Category.objects.create(code="OBT", name="Obat", sort_order=1)
		self.facility = Facility.objects.create(
			code="PKM-CON-01",
			name="Puskesmas Pemakaian",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.other_facility = Facility.objects.create(
			code="PKM-CON-02",
			name="Puskesmas Pemakaian Lain",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.item = Item.objects.create(
			nama_barang="Amoxicillin 500 mg",
			satuan=self.unit,
			kategori=self.category,
			is_active=True,
		)
		self.subunit = PuskesmasSubunit.objects.create(
			facility=self.facility,
			name="Poli Umum",
			subunit_type=PuskesmasSubunit.SubunitType.TREATMENT_ROOM,
		)
		self.other_subunit = PuskesmasSubunit.objects.create(
			facility=self.other_facility,
			name="Pustu A",
			subunit_type=PuskesmasSubunit.SubunitType.HELPER_SITE,
		)
		self.operator = User.objects.create_user(
			username="operator-consumption-view",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=self.facility,
		)
		self.other_operator = User.objects.create_user(
			username="operator-consumption-other",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=self.other_facility,
		)
		self.admin = User.objects.create_superuser(
			username="admin-consumption",
			email="admin-consumption@example.com",
			password="TestPassword123!",
		)

	def _create_lplpo(self, status=LPLPO.Status.DRAFT):
		lplpo = LPLPO.objects.create(
			facility=self.facility,
			bulan=2,
			tahun=2026,
			status=status,
			created_by=self.operator,
		)
		LPLPOItem.objects.create(
			lplpo=lplpo,
			item=self.item,
			stock_awal=Decimal("10.00"),
			penerimaan=Decimal("5.00"),
			harga_satuan=Decimal("1000.00"),
		)
		return lplpo

	def _payload(self, quantity="7"):
		return {
			"facility": str(self.facility.pk),
			"bulan": "2",
			"tahun": "2026",
			"notes": "  Catatan konsumsi  ",
			f"qty_{self.item.pk}_{self.subunit.pk}": quantity,
		}

	def test_create_consumption_syncs_editable_lplpo(self):
		lplpo = self._create_lplpo()
		self.client.force_login(self.operator)

		response = self.client.post(
			reverse("puskesmas:consumption_create"),
			self._payload(quantity="7"),
		)

		self.assertEqual(response.status_code, 302)
		consumption = PuskesmasConsumption.objects.get()
		self.assertEqual(consumption.notes, "Catatan konsumsi")
		line = lplpo.items.get(item=self.item)
		self.assertEqual(line.pemakaian, Decimal("7.00"))
		self.assertEqual(line.stock_keseluruhan, Decimal("8.00"))

	def test_create_consumption_is_blocked_when_lplpo_submitted(self):
		self._create_lplpo(status=LPLPO.Status.SUBMITTED)
		self.client.force_login(self.operator)

		response = self.client.post(
			reverse("puskesmas:consumption_create"),
			self._payload(quantity="7"),
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Pemakaian untuk periode ini tidak dapat diubah")
		self.assertEqual(PuskesmasConsumption.objects.count(), 0)

	def test_cross_facility_consumption_detail_gets_403(self):
		consumption = PuskesmasConsumption.objects.create(
			facility=self.other_facility,
			bulan=2,
			tahun=2026,
			created_by=self.other_operator,
		)
		PuskesmasConsumptionEntry.objects.create(
			consumption=consumption,
			item=self.item,
			subunit=self.other_subunit,
			quantity=3,
		)

		self.client.force_login(self.operator)
		response = self.client.get(reverse("puskesmas:consumption_detail", args=[consumption.pk]))

		self.assertEqual(response.status_code, 403)

	@override_settings(PUSKESMAS_CONSUMPTION_MUTATION_RATE_LIMIT="1/m", RATELIMIT_USE_CACHE="locmem")
	def test_create_consumption_returns_429_when_rate_limited(self):
		self.client.force_login(self.operator)
		first_response = self.client.post(
			reverse("puskesmas:consumption_create"),
			self._payload(quantity="1"),
		)
		second_response = self.client.post(
			reverse("puskesmas:consumption_create"),
			self._payload(quantity="2"),
		)

		self.assertEqual(first_response.status_code, 302)
		self.assertEqual(second_response.status_code, 429)
		self.assertContains(
			second_response,
			"Terlalu banyak percobaan pada aksi ini",
			status_code=429,
		)

	def test_superuser_load_matrix_does_not_create_blank_consumption(self):
		self.client.force_login(self.admin)

		response = self.client.post(
			reverse("puskesmas:consumption_create"),
			{
				"facility": str(self.facility.pk),
				"bulan": "2",
				"tahun": "2026",
				"notes": "",
				"action": "load_matrix",
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(PuskesmasConsumption.objects.count(), 0)
		self.assertContains(response, self.subunit.name)
		self.assertContains(
			response,
			f'qty_{self.item.pk}_{self.subunit.pk}',
			html=False,
		)

	def test_edit_preserves_existing_inactive_item_entry(self):
		lplpo = self._create_lplpo()
		inactive_item = Item.objects.create(
			nama_barang="Item Nonaktif",
			satuan=self.unit,
			kategori=self.category,
			is_active=True,
		)
		consumption = PuskesmasConsumption.objects.create(
			facility=self.facility,
			bulan=2,
			tahun=2026,
			notes="Catatan awal",
			created_by=self.operator,
			updated_by=self.operator,
		)
		PuskesmasConsumptionEntry.objects.create(
			consumption=consumption,
			item=self.item,
			subunit=self.subunit,
			quantity=7,
		)
		PuskesmasConsumptionEntry.objects.create(
			consumption=consumption,
			item=inactive_item,
			subunit=self.subunit,
			quantity=4,
		)
		inactive_item.is_active = False
		inactive_item.save(update_fields=["is_active", "updated_at"])
		LPLPOItem.objects.create(
			lplpo=lplpo,
			item=inactive_item,
			stock_awal=Decimal("3.00"),
			penerimaan=Decimal("0.00"),
			harga_satuan=Decimal("500.00"),
		)

		self.client.force_login(self.operator)
		response = self.client.post(
			reverse("puskesmas:consumption_edit", args=[consumption.pk]),
			{
				"facility": str(self.facility.pk),
				"bulan": "2",
				"tahun": "2026",
				"notes": "Catatan diperbarui",
				f"qty_{self.item.pk}_{self.subunit.pk}": "8",
				f"qty_{inactive_item.pk}_{self.subunit.pk}": "4",
			},
		)

		self.assertEqual(response.status_code, 302)
		consumption.refresh_from_db()
		self.assertEqual(consumption.notes, "Catatan diperbarui")
		self.assertTrue(
			consumption.entries.filter(
				item=inactive_item,
				subunit=self.subunit,
				quantity=4,
			).exists()
		)
		self.assertEqual(lplpo.items.get(item=self.item).pemakaian, Decimal("8.00"))
		self.assertEqual(
			lplpo.items.get(item=inactive_item).pemakaian,
			Decimal("4.00"),
		)

	def test_detail_shows_existing_inactive_item_entry(self):
		inactive_item = Item.objects.create(
			nama_barang="Vitamin C",
			satuan=self.unit,
			kategori=self.category,
			is_active=True,
		)
		consumption = PuskesmasConsumption.objects.create(
			facility=self.facility,
			bulan=2,
			tahun=2026,
			created_by=self.operator,
		)
		PuskesmasConsumptionEntry.objects.create(
			consumption=consumption,
			item=inactive_item,
			subunit=self.subunit,
			quantity=5,
		)
		inactive_item.is_active = False
		inactive_item.save(update_fields=["is_active", "updated_at"])

		self.client.force_login(self.operator)
		response = self.client.get(reverse("puskesmas:consumption_detail", args=[consumption.pk]))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, inactive_item.nama_barang)


class PuskesmasReportViewTests(SecureClientDefaultsMixin, TestCase):
	"""Tests for the Puskesmas report views and facility isolation rules."""

	def setUp(self):
		super().setUp()
		self.unit = Unit.objects.create(code="TAB", name="Tablet")
		self.category = Category.objects.create(code="OBT", name="Obat", sort_order=1)
		self.facility = Facility.objects.create(
			code="PKM-01",
			name="Puskesmas Satu",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.other_facility = Facility.objects.create(
			code="PKM-02",
			name="Puskesmas Dua",
			facility_type=Facility.FacilityType.PUSKESMAS,
		)
		self.item = Item.objects.create(
			nama_barang="Amoxicillin 500 mg",
			satuan=self.unit,
			kategori=self.category,
			is_active=True,
		)
		# Puskesmas operator linked to facility
		self.operator = User.objects.create_user(
			username="operator-report",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=self.facility,
		)
		self.report_operator = User.objects.create_user(
			username="operator-report-enabled",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=self.facility,
		)
		ensure_default_module_access(self.report_operator, overwrite=True)
		ModuleAccess.objects.update_or_create(
			user=self.report_operator,
			module=ModuleAccess.Module.REPORTS,
			defaults={"scope": ModuleAccess.Scope.VIEW},
		)
		# Operator from a different facility
		self.other_operator = User.objects.create_user(
			username="operator-other-report",
			password="TestPassword123!",
			role=User.Role.PUSKESMAS,
			facility=self.other_facility,
		)
		# Super admin
		self.admin = User.objects.create_superuser(
			username="admin-report",
			email="admin-report@example.com",
			password="TestPassword123!",
		)
		self.staff_with_facility = User.objects.create_user(
			username="gudang-report",
			password="TestPassword123!",
			role=User.Role.GUDANG,
			facility=self.facility,
		)
		ensure_default_module_access(self.staff_with_facility, overwrite=True)
		self.staff_without_facility = User.objects.create_user(
			username="gudang-no-facility",
			password="TestPassword123!",
			role=User.Role.GUDANG,
		)
		ensure_default_module_access(self.staff_without_facility, overwrite=True)
		# Non-puskesmas user without puskesmas perm (AUDITOR with NONE scope)
		self.auditor = User.objects.create_user(
			username="auditor-report",
			password="TestPassword123!",
			role=User.Role.AUDITOR,
		)
		ModuleAccess.objects.update_or_create(
			user=self.auditor,
			module=ModuleAccess.Module.PUSKESMAS,
			defaults={"scope": ModuleAccess.Scope.NONE},
		)

	# ────────────── Permission / Auth checks ──────────────

	def test_penerimaan_requires_login(self):
		response = self.client.get(reverse("puskesmas:report_penerimaan"))
		self.assertEqual(response.status_code, 302)
		self.assertIn("/login/", response["Location"])

	def test_pemakaian_requires_login(self):
		response = self.client.get(reverse("puskesmas:report_pemakaian"))
		self.assertEqual(response.status_code, 302)
		self.assertIn("/login/", response["Location"])

	def test_persediaan_requires_login(self):
		response = self.client.get(reverse("puskesmas:report_persediaan"))
		self.assertEqual(response.status_code, 302)
		self.assertIn("/login/", response["Location"])

	def test_auditor_with_none_scope_cannot_access_penerimaan(self):
		self.client.force_login(self.auditor)
		response = self.client.get(reverse("puskesmas:report_penerimaan"))
		self.assertEqual(response.status_code, 403)

	def test_auditor_with_none_scope_cannot_access_pemakaian(self):
		self.client.force_login(self.auditor)
		response = self.client.get(reverse("puskesmas:report_pemakaian"))
		self.assertEqual(response.status_code, 403)

	def test_auditor_with_none_scope_cannot_access_persediaan(self):
		self.client.force_login(self.auditor)
		response = self.client.get(reverse("puskesmas:report_persediaan"))
		self.assertEqual(response.status_code, 403)

	def test_puskesmas_operator_without_reports_scope_cannot_access_penerimaan(self):
		self.client.force_login(self.operator)
		response = self.client.get(reverse("puskesmas:report_penerimaan"))
		self.assertEqual(response.status_code, 403)

	def test_puskesmas_operator_without_reports_scope_cannot_access_pemakaian(self):
		self.client.force_login(self.operator)
		response = self.client.get(reverse("puskesmas:report_pemakaian"))
		self.assertEqual(response.status_code, 403)

	def test_puskesmas_operator_without_reports_scope_cannot_access_persediaan(self):
		self.client.force_login(self.operator)
		response = self.client.get(reverse("puskesmas:report_persediaan"))
		self.assertEqual(response.status_code, 403)

	def test_report_enabled_puskesmas_operator_can_access_all_four_reports(self):
		self.client.force_login(self.report_operator)
		for url_name in (
			"report_penerimaan",
			"report_pemakaian",
			"report_persediaan",
			"report_rekap_persediaan",
		):
			with self.subTest(url_name=url_name):
				response = self.client.get(reverse(f"puskesmas:{url_name}"))
				self.assertEqual(response.status_code, 200)

	def test_puskesmas_report_sidebar_requires_reports_scope(self):
		self.client.force_login(self.operator)
		response = self.client.get(reverse("puskesmas:request_list"))
		self.assertEqual(response.status_code, 200)
		self.assertNotContains(response, "Laporan Puskesmas")

		self.client.force_login(self.report_operator)
		response = self.client.get(reverse("puskesmas:request_list"))
		self.assertEqual(response.status_code, 200)
		self.assertContains(response, "Laporan Puskesmas")

	def test_report_page_does_not_show_request_back_link(self):
		self.client.force_login(self.report_operator)

		response = self.client.get(reverse("puskesmas:report_penerimaan"))

		self.assertEqual(response.status_code, 200)
		self.assertNotContains(response, 'title="Kembali ke Permintaan Puskesmas"', html=False)

	def test_admin_can_access_all_three_reports(self):
		self.client.force_login(self.admin)
		for url_name in (
			"report_penerimaan",
			"report_pemakaian",
			"report_persediaan",
			"report_rekap_persediaan",
		):
			with self.subTest(url_name=url_name):
				response = self.client.get(reverse(f"puskesmas:{url_name}"))
				self.assertEqual(response.status_code, 200)

	def test_non_superuser_report_scope_defaults_to_linked_facility(self):
		from datetime import date as dt

		sbbk_own = self._make_sbbk(self.facility, received_date=dt(2026, 3, 15))
		PuskesmasSBBKItem.objects.create(sbbk=sbbk_own, item=self.item, quantity=10, unit_price=1000)
		sbbk_other = self._make_sbbk(self.other_facility, received_date=dt(2026, 3, 15))
		PuskesmasSBBKItem.objects.create(sbbk=sbbk_other, item=self.item, quantity=5, unit_price=1000)

		self.client.force_login(self.report_operator)
		response = self.client.get(
			reverse("puskesmas:report_penerimaan"),
			{"start_date": "2026-03-01", "end_date": "2026-03-31"},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(response.context["selected_facility_name"], self.facility.name)
		document_numbers = [row["document_number"] for row in response.context["report_data"]]
		self.assertEqual(document_numbers, [sbbk_own.document_number])

	def test_non_superuser_report_scope_ignores_mismatched_facility_query(self):
		from datetime import date as dt

		sbbk_own = self._make_sbbk(self.facility, received_date=dt(2026, 3, 15))
		PuskesmasSBBKItem.objects.create(sbbk=sbbk_own, item=self.item, quantity=10, unit_price=1000)
		sbbk_other = self._make_sbbk(self.other_facility, received_date=dt(2026, 3, 15))
		PuskesmasSBBKItem.objects.create(sbbk=sbbk_other, item=self.item, quantity=5, unit_price=1000)

		self.client.force_login(self.report_operator)
		response = self.client.get(
			reverse("puskesmas:report_penerimaan"),
			{
				"facility": str(self.other_facility.pk),
				"start_date": "2026-03-01",
				"end_date": "2026-03-31",
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(response.context["selected_facility_name"], self.facility.name)
		document_numbers = [row["document_number"] for row in response.context["report_data"]]
		self.assertEqual(document_numbers, [sbbk_own.document_number])

	def test_non_superuser_without_linked_facility_gets_403_on_all_reports(self):
		self.client.force_login(self.staff_without_facility)

		for url_name in (
			"report_penerimaan",
			"report_pemakaian",
			"report_persediaan",
			"report_rekap_persediaan",
		):
			with self.subTest(url_name=url_name):
				response = self.client.get(reverse(f"puskesmas:{url_name}"))
				self.assertEqual(response.status_code, 403)

	def test_report_endpoints_deny_excel_exports_without_reports_scope(self):
		self.client.force_login(self.operator)

		for url_name, query in (
			("report_penerimaan", {"format": "excel"}),
			("report_pemakaian", {"format": "excel"}),
			("report_persediaan", {"format": "excel"}),
			("report_rekap_persediaan", {"format": "excel"}),
		):
			with self.subTest(url_name=url_name):
				response = self.client.get(reverse(f"puskesmas:{url_name}"), query)
				self.assertEqual(response.status_code, 403)

	# ────────────── Penerimaan data correctness ──────────────

	def _make_distribution(self, facility, status, dist_type=None, distributed_date=None):
		from datetime import date as dt
		dist = Distribution.objects.create(
			distribution_type=dist_type or Distribution.DistributionType.LPLPO,
			facility=facility,
			request_date=distributed_date or dt(2026, 3, 15),
			status=status,
			distributed_date=distributed_date or dt(2026, 3, 15),
			created_by=self.admin,
		)
		return dist

	def _make_sbbk(self, facility, received_date=None):
		from datetime import date as dt
		return PuskesmasSBBK.objects.create(
			facility=facility,
			received_date=received_date or dt(2026, 3, 15),
			created_by=self.admin,
		)

	def test_penerimaan_only_shows_distributed_status(self):
		from datetime import date as dt

		sbbk_ok = self._make_sbbk(self.facility, received_date=dt(2026, 3, 15))
		PuskesmasSBBKItem.objects.create(sbbk=sbbk_ok, item=self.item, quantity=10, unit_price=1000)
		sbbk_outside = self._make_sbbk(self.facility, received_date=dt(2026, 2, 10))
		PuskesmasSBBKItem.objects.create(sbbk=sbbk_outside, item=self.item, quantity=5, unit_price=1000)

		self.client.force_login(self.report_operator)
		response = self.client.get(
			reverse("puskesmas:report_penerimaan"),
			{"start_date": "2026-03-01", "end_date": "2026-03-31"},
		)
		self.assertEqual(response.status_code, 200)
		report_data = response.context["report_data"]
		self.assertEqual(len(report_data), 1)
		self.assertEqual(report_data[0]["document_number"], sbbk_ok.document_number)

	def test_penerimaan_isolates_facility(self):
		from datetime import date as dt

		sbbk_own = self._make_sbbk(self.facility, received_date=dt(2026, 3, 15))
		PuskesmasSBBKItem.objects.create(sbbk=sbbk_own, item=self.item, quantity=10, unit_price=1000)
		sbbk_other = self._make_sbbk(self.other_facility, received_date=dt(2026, 3, 15))
		PuskesmasSBBKItem.objects.create(sbbk=sbbk_other, item=self.item, quantity=5, unit_price=1000)

		self.client.force_login(self.report_operator)
		response = self.client.get(
			reverse("puskesmas:report_penerimaan"),
			{"start_date": "2026-03-01", "end_date": "2026-03-31"},
		)
		self.assertEqual(response.status_code, 200)
		report_data = response.context["report_data"]
		document_numbers = [r["document_number"] for r in report_data]
		self.assertIn(sbbk_own.document_number, document_numbers)
		self.assertNotIn(sbbk_other.document_number, document_numbers)

	# ────────────── Pemakaian data correctness ──────────────

	def _make_lplpo(self, facility, bulan, tahun, status):
		from apps.lplpo.models import LPLPO
		lplpo = LPLPO.objects.create(
			facility=facility,
			bulan=bulan,
			tahun=tahun,
			status=status,
			created_by=self.admin,
		)
		return lplpo

	def test_pemakaian_only_shows_finalized_lplpo(self):
		from apps.lplpo.models import LPLPO, LPLPOItem

		# CLOSED LPLPO — should appear
		lplpo_ok = self._make_lplpo(self.facility, 3, 2026, LPLPO.Status.CLOSED)
		LPLPOItem.objects.create(
			lplpo=lplpo_ok, item=self.item,
			pemakaian=100, stock_awal=50, penerimaan=80,
			stock_keseluruhan=30, permintaan_jumlah=50,
		)

		# SUBMITTED LPLPO — should NOT appear
		lplpo_bad = self._make_lplpo(self.facility, 4, 2026, LPLPO.Status.SUBMITTED)
		LPLPOItem.objects.create(
			lplpo=lplpo_bad, item=self.item,
			pemakaian=20, stock_awal=30,
		)

		self.client.force_login(self.report_operator)
		response = self.client.get(
			reverse("puskesmas:report_pemakaian"),
			{"year": "2026"},
		)
		self.assertEqual(response.status_code, 200)
		report_data = response.context["report_data"]
		periods = [r["period_display"] for r in report_data]
		self.assertIn("Maret 2026", periods)
		self.assertNotIn("April 2026", periods)

	def test_pemakaian_isolates_facility(self):
		from apps.lplpo.models import LPLPO, LPLPOItem

		lplpo_own = self._make_lplpo(self.facility, 3, 2026, LPLPO.Status.CLOSED)
		LPLPOItem.objects.create(
			lplpo=lplpo_own, item=self.item,
			pemakaian=100, stock_awal=50,
		)
		lplpo_other = self._make_lplpo(self.other_facility, 3, 2026, LPLPO.Status.CLOSED)
		LPLPOItem.objects.create(
			lplpo=lplpo_other, item=self.item,
			pemakaian=200, stock_awal=100,
		)

		self.client.force_login(self.report_operator)
		response = self.client.get(
			reverse("puskesmas:report_pemakaian"),
			{"year": "2026"},
		)
		self.assertEqual(response.status_code, 200)
		report_data = response.context["report_data"]
		# Each item must belong only to operator's facility
		for row in report_data:
			self.assertEqual(row["period_display"], "Maret 2026")
		# The data from other facility should not appear (only one LPLPO for our facility)
		self.assertEqual(len(report_data), 1)

	def test_pemakaian_distributed_status_included(self):
		from apps.lplpo.models import LPLPO, LPLPOItem

		lplpo = self._make_lplpo(self.facility, 2, 2026, LPLPO.Status.DISTRIBUTED)
		LPLPOItem.objects.create(
			lplpo=lplpo, item=self.item,
			pemakaian=60, stock_awal=80,
		)

		self.client.force_login(self.report_operator)
		response = self.client.get(
			reverse("puskesmas:report_pemakaian"),
			{"year": "2026"},
		)
		self.assertEqual(response.status_code, 200)
		report_data = response.context["report_data"]
		periods = [r["period_display"] for r in report_data]
		self.assertIn("Februari 2026", periods)

	# ────────────── Persediaan LPLPO-based check ──────────────

	def test_persediaan_returns_200_with_form_for_operator(self):
		self.client.force_login(self.report_operator)
		response = self.client.get(reverse("puskesmas:report_persediaan"))
		self.assertEqual(response.status_code, 200)
		self.assertIn("form", response.context)

	def test_persediaan_returns_200_with_form_for_admin(self):
		self.client.force_login(self.admin)
		response = self.client.get(reverse("puskesmas:report_persediaan"))
		self.assertEqual(response.status_code, 200)
		self.assertIn("form", response.context)

	def test_persediaan_isolates_facility_via_lplpo(self):
		"""Operator sees only their facility's LPLPO stock, not other facilities."""
		from apps.items.models import Item, Unit, Category
		from apps.lplpo.models import LPLPO, LPLPOItem

		lplpo_own = LPLPO.objects.create(
			facility=self.facility,
			bulan=4,
			tahun=2026,
			status=LPLPO.Status.CLOSED,
			created_by=self.admin,
		)
		LPLPOItem.objects.create(
			lplpo=lplpo_own,
			item=self.item,
			stock_awal=50,
			penerimaan=20,
			pemakaian=10,
		)

		lplpo_other = LPLPO.objects.create(
			facility=self.other_facility,
			bulan=4,
			tahun=2026,
			status=LPLPO.Status.CLOSED,
			created_by=self.admin,
		)
		LPLPOItem.objects.create(
			lplpo=lplpo_other,
			item=self.item,
			stock_awal=999,
			penerimaan=999,
			pemakaian=0,
		)

		self.client.force_login(self.report_operator)
		response = self.client.get(
			reverse("puskesmas:report_persediaan"),
			{"year": "2026", "period": "q2"},
			follow=True,
		)
		self.assertEqual(response.status_code, 200)
		report_data = response.context["report_data"]
		# Only one item row, from operator's facility LPLPO (stock_keseluruhan = 50+20-10=60)
		self.assertEqual(len(report_data), 1)
		self.assertEqual(report_data[0]["stock_keseluruhan"], 60)
		self.assertEqual(response.context["period_label"], "Triwulan II")

	def test_persediaan_admin_all_facilities_aggregates_same_item_rows(self):
		from apps.lplpo.models import LPLPO, LPLPOItem

		lplpo_own = LPLPO.objects.create(
			facility=self.facility,
			bulan=4,
			tahun=2026,
			status=LPLPO.Status.CLOSED,
			created_by=self.admin,
		)
		LPLPOItem.objects.create(
			lplpo=lplpo_own,
			item=self.item,
			stock_awal=50,
			penerimaan=20,
			pemakaian=10,
		)

		lplpo_other = LPLPO.objects.create(
			facility=self.other_facility,
			bulan=4,
			tahun=2026,
			status=LPLPO.Status.CLOSED,
			created_by=self.admin,
		)
		LPLPOItem.objects.create(
			lplpo=lplpo_other,
			item=self.item,
			stock_awal=30,
			penerimaan=5,
			pemakaian=3,
		)

		self.client.force_login(self.admin)
		response = self.client.get(
			reverse("puskesmas:report_persediaan"),
			{"year": "2026", "period": "q2"},
		)

		self.assertEqual(response.status_code, 200)
		report_data = response.context["report_data"]
		self.assertEqual(len(report_data), 1)
		self.assertEqual(report_data[0]["nama_barang"], self.item.nama_barang)
		self.assertEqual(report_data[0]["stock_keseluruhan"], 92)

	def test_persediaan_period_filter_uses_period_end_month(self):
		from apps.lplpo.models import LPLPO, LPLPOItem

		march = LPLPO.objects.create(
			facility=self.facility,
			bulan=3,
			tahun=2026,
			status=LPLPO.Status.CLOSED,
			created_by=self.admin,
		)
		LPLPOItem.objects.create(
			lplpo=march,
			item=self.item,
			stock_awal=10,
			penerimaan=5,
			pemakaian=2,
		)

		june = LPLPO.objects.create(
			facility=self.facility,
			bulan=6,
			tahun=2026,
			status=LPLPO.Status.CLOSED,
			created_by=self.admin,
		)
		LPLPOItem.objects.create(
			lplpo=june,
			item=self.item,
			stock_awal=13,
			penerimaan=4,
			pemakaian=3,
		)

		self.client.force_login(self.report_operator)
		response = self.client.get(
			reverse("puskesmas:report_persediaan"),
			{"year": "2026", "period": "q1"},
			follow=True,
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(response.context["period_label"], "Triwulan I")
		self.assertEqual(response.context["report_data"][0]["stock_keseluruhan"], 13)

	def test_persediaan_uses_confirmed_receipts_and_detailed_consumption_after_lplpo_baseline(self):
		from datetime import date as dt
		from apps.lplpo.models import LPLPO, LPLPOItem

		baseline = LPLPO.objects.create(
			facility=self.facility,
			bulan=3,
			tahun=2026,
			status=LPLPO.Status.CLOSED,
			created_by=self.admin,
		)
		LPLPOItem.objects.create(
			lplpo=baseline,
			item=self.item,
			stock_awal=20,
			penerimaan=0,
			pemakaian=5,
		)
		april_receipt = PuskesmasSBBK.objects.create(
			facility=self.facility,
			received_date=dt(2026, 4, 10),
			status=PuskesmasSBBK.ReceiptStatus.CONFIRMED,
			created_by=self.admin,
		)
		PuskesmasSBBKItem.objects.create(
			sbbk=april_receipt,
			item=self.item,
			quantity=Decimal("7"),
			unit_price=Decimal("1000"),
			batch_lot="PKM-RCV-001",
		)
		draft_receipt = PuskesmasSBBK.objects.create(
			facility=self.facility,
			received_date=dt(2026, 5, 10),
			status=PuskesmasSBBK.ReceiptStatus.DRAFT,
			created_by=self.admin,
		)
		PuskesmasSBBKItem.objects.create(
			sbbk=draft_receipt,
			item=self.item,
			quantity=Decimal("99"),
			unit_price=Decimal("1000"),
			batch_lot="PKM-DRAFT",
		)
		subunit = PuskesmasSubunit.objects.create(
			facility=self.facility,
			name="Poli Umum",
			subunit_type=PuskesmasSubunit.SubunitType.TREATMENT_ROOM,
		)
		may_consumption = PuskesmasConsumption.objects.create(
			facility=self.facility,
			bulan=5,
			tahun=2026,
			created_by=self.admin,
		)
		PuskesmasConsumptionEntry.objects.create(
			consumption=may_consumption,
			item=self.item,
			subunit=subunit,
			quantity=4,
		)
		baseline_month_consumption = PuskesmasConsumption.objects.create(
			facility=self.facility,
			bulan=3,
			tahun=2026,
			created_by=self.admin,
		)
		PuskesmasConsumptionEntry.objects.create(
			consumption=baseline_month_consumption,
			item=self.item,
			subunit=subunit,
			quantity=50,
		)

		self.client.force_login(self.report_operator)
		response = self.client.get(
			reverse("puskesmas:report_persediaan"),
			{"year": "2026", "period": "s1"},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(response.context["period_label"], "Semester I")
		self.assertEqual(len(response.context["report_data"]), 1)
		self.assertEqual(response.context["report_data"][0]["stock_keseluruhan"], 18)

	def test_persediaan_does_not_use_distribution_fallback_without_lplpo_baseline(self):
		from datetime import date as dt

		distribution = self._make_distribution(
			self.facility,
			Distribution.Status.DISTRIBUTED,
			distributed_date=dt(2026, 4, 10),
		)
		DistributionItem.objects.create(
			distribution=distribution,
			item=self.item,
			quantity_requested=Decimal("12"),
			quantity_approved=Decimal("12"),
		)

		self.client.force_login(self.report_operator)
		response = self.client.get(
			reverse("puskesmas:report_persediaan"),
			{"year": "2026", "period": "q2"},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(response.context["report_data"], [])

	def test_rekap_persediaan_aggregates_asset_valuation_by_category(self):
		from apps.lplpo.models import LPLPO, LPLPOItem
		other_category = Category.objects.create(
			code="RGN",
			name="Reagen",
			sort_order=2,
		)
		item_same_category = Item.objects.create(
			nama_barang="Paracetamol 500 mg",
			satuan=self.unit,
			kategori=self.category,
			is_active=True,
		)
		item_other_category = Item.objects.create(
			nama_barang="Reagen A",
			satuan=self.unit,
			kategori=other_category,
			is_active=True,
		)

		january = LPLPO.objects.create(
			facility=self.facility,
			bulan=1,
			tahun=2026,
			status=LPLPO.Status.CLOSED,
			created_by=self.admin,
		)
		LPLPOItem.objects.create(
			lplpo=january,
			item=self.item,
			stock_awal=10,
			penerimaan=0,
			harga_satuan=Decimal("1000.00"),
			pemakaian=4,
		)
		LPLPOItem.objects.create(
			lplpo=january,
			item=item_same_category,
			stock_awal=5,
			penerimaan=0,
			harga_satuan=Decimal("2000.00"),
			pemakaian=1,
		)
		LPLPOItem.objects.create(
			lplpo=january,
			item=item_other_category,
			stock_awal=3,
			penerimaan=1,
			harga_satuan=Decimal("500.00"),
			pemakaian=1,
		)

		february = LPLPO.objects.create(
			facility=self.facility,
			bulan=2,
			tahun=2026,
			status=LPLPO.Status.CLOSED,
			created_by=self.admin,
		)
		LPLPOItem.objects.create(
			lplpo=february,
			item=self.item,
			stock_awal=6,
			penerimaan=5,
			harga_satuan=Decimal("1200.00"),
			pemakaian=3,
		)
		LPLPOItem.objects.create(
			lplpo=february,
			item=item_same_category,
			stock_awal=4,
			penerimaan=2,
			harga_satuan=Decimal("2500.00"),
			pemakaian=1,
		)
		LPLPOItem.objects.create(
			lplpo=february,
			item=item_other_category,
			stock_awal=3,
			penerimaan=2,
			harga_satuan=Decimal("700.00"),
			pemakaian=1,
		)

		self.client.force_login(self.report_operator)
		response = self.client.get(
			reverse("puskesmas:report_rekap_persediaan"),
			{"year": "2026", "period": "q1"},
			follow=True,
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(len(response.context["rekap_data"]), 2)
		obat_row = response.context["rekap_data"][0]
		reagen_row = response.context["rekap_data"][1]
		self.assertEqual(obat_row["kategori"], self.category.name)
		self.assertEqual(obat_row["saldo_awal"], Decimal("20000.00"))
		self.assertEqual(obat_row["nilai_terima"], Decimal("11000.00"))
		self.assertEqual(obat_row["nilai_keluar"], Decimal("12100.00"))
		self.assertEqual(obat_row["saldo_akhir"], Decimal("22100.00"))
		self.assertEqual(reagen_row["kategori"], other_category.name)
		self.assertEqual(reagen_row["saldo_awal"], Decimal("1500.00"))
		self.assertEqual(reagen_row["nilai_terima"], Decimal("1900.00"))
		self.assertEqual(reagen_row["nilai_keluar"], Decimal("1200.00"))
		self.assertEqual(reagen_row["saldo_akhir"], Decimal("2800.00"))
		self.assertEqual(response.context["totals"]["saldo_awal"], Decimal("21500.00"))
		self.assertEqual(response.context["totals"]["nilai_terima"], Decimal("12900.00"))
		self.assertEqual(response.context["totals"]["nilai_keluar"], Decimal("13300.00"))
		self.assertEqual(response.context["totals"]["saldo_akhir"], Decimal("24900.00"))
		self.assertContains(response, "URAIAN")
		self.assertContains(response, "SALDO AWAL 2026")
		self.assertContains(response, "Rp 22.100,00")
		self.assertContains(response, "Triwulan I")

	def test_rekap_persediaan_admin_aggregates_across_facilities(self):
		from apps.lplpo.models import LPLPO, LPLPOItem

		lplpo_own = LPLPO.objects.create(
			facility=self.facility,
			bulan=1,
			tahun=2026,
			status=LPLPO.Status.CLOSED,
			created_by=self.admin,
		)
		LPLPOItem.objects.create(
			lplpo=lplpo_own,
			item=self.item,
			stock_awal=10,
			penerimaan=5,
			harga_satuan=Decimal("100.00"),
			pemakaian=2,
		)
		lplpo_other = LPLPO.objects.create(
			facility=self.other_facility,
			bulan=1,
			tahun=2026,
			status=LPLPO.Status.CLOSED,
			created_by=self.admin,
		)
		LPLPOItem.objects.create(
			lplpo=lplpo_other,
			item=self.item,
			stock_awal=20,
			penerimaan=1,
			harga_satuan=Decimal("200.00"),
			pemakaian=5,
		)

		self.client.force_login(self.admin)
		response = self.client.get(
			reverse("puskesmas:report_rekap_persediaan"),
			{"year": "2026", "period": "q1"},
			follow=True,
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(len(response.context["rekap_data"]), 1)
		row = response.context["rekap_data"][0]
		self.assertEqual(row["kategori"], self.category.name)
		self.assertEqual(row["saldo_awal"], Decimal("5000.00"))
		self.assertEqual(row["nilai_terima"], Decimal("700.00"))
		self.assertEqual(row["nilai_keluar"], Decimal("1200.00"))
		self.assertEqual(row["saldo_akhir"], Decimal("4500.00"))
		self.assertEqual(response.context["totals"]["saldo_awal"], Decimal("5000.00"))
		self.assertEqual(response.context["totals"]["nilai_terima"], Decimal("700.00"))
		self.assertEqual(response.context["totals"]["nilai_keluar"], Decimal("1200.00"))
		self.assertEqual(response.context["totals"]["saldo_akhir"], Decimal("4500.00"))

	def test_rekap_persediaan_preserves_zero_category_sort_order(self):
		zero_category = Category.objects.create(
			code="ALK",
			name="Alkes",
			sort_order=0,
		)
		zero_item = Item.objects.create(
			nama_barang="Masker Bedah",
			satuan=self.unit,
			kategori=zero_category,
			is_active=True,
		)
		lplpo = LPLPO.objects.create(
			facility=self.facility,
			bulan=1,
			tahun=2026,
			status=LPLPO.Status.CLOSED,
			created_by=self.admin,
		)
		LPLPOItem.objects.create(
			lplpo=lplpo,
			item=zero_item,
			stock_awal=1,
			harga_satuan=Decimal("100.00"),
		)
		LPLPOItem.objects.create(
			lplpo=lplpo,
			item=self.item,
			stock_awal=1,
			harga_satuan=Decimal("100.00"),
		)

		self.client.force_login(self.report_operator)
		response = self.client.get(
			reverse("puskesmas:report_rekap_persediaan"),
			{"year": "2026", "period": "q1"},
			follow=True,
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(
			[row["kategori"] for row in response.context["rekap_data"]],
			[zero_category.name, self.category.name],
		)

	def test_rekap_persediaan_excel_export_uses_category_summary_headers(self):
		from apps.puskesmas.exports import export_puskesmas_rekap_persediaan_excel

		response = export_puskesmas_rekap_persediaan_excel(
			rekap_data=[
				{
					"kategori": self.category.name,
					"saldo_awal": Decimal("10000.00"),
					"nilai_terima": Decimal("6000.00"),
					"nilai_keluar": Decimal("7600.00"),
					"saldo_akhir": Decimal("9600.00"),
				}
			],
			totals={
				"saldo_awal": Decimal("10000.00"),
				"nilai_terima": Decimal("6000.00"),
				"nilai_keluar": Decimal("7600.00"),
				"saldo_akhir": Decimal("9600.00"),
			},
			year=2026,
			period_label="Triwulan I",
			facility_name=self.facility.name,
		)

		self.assertEqual(
			response["Content-Type"],
			"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		)
		workbook = load_workbook(BytesIO(response.content))
		sheet = workbook.active
		headers = [sheet.cell(row=4, column=idx).value for idx in range(1, 7)]
		self.assertEqual(
			headers,
			[
				"No",
				"Uraian",
				"Saldo Awal 2026\n(Rp)",
				"Nilai Terima 2026\n(Rp)",
				"Nilai Keluar 2026\n(Rp)",
				"Saldo Akhir 2026\n(Rp)",
			],
		)
		self.assertEqual(sheet.cell(row=5, column=2).value, self.category.name)
		self.assertEqual(sheet.cell(row=6, column=2).value, "TOTAL")

	def test_penerimaan_excel_neutralizes_formula_prefixed_text_and_keeps_numeric_cells(self):
		response = export_puskesmas_penerimaan_excel(
			[
				{
					"received_date": None,
					"document_number": "=DIST-001",
					"distribution_document_number": "@DIST-SOURCE",
					"nama_barang": "@Amoxicillin 500 mg",
					"satuan": "-Tablet",
					"quantity": Decimal("5"),
					"unit_price": Decimal("2500"),
					"batch_lot": "+BATCH-LOT",
					"expiry_date": None,
					"notes": "=BATCH-01",
				}
			],
			"2026-03-01",
			"2026-03-31",
			"=Puskesmas Formula",
		)

		workbook = load_workbook(BytesIO(response.content))
		sheet = workbook.active

		self.assertEqual(
			sheet["A2"].value,
			"Fasilitas: =Puskesmas Formula | Periode: 2026-03-01 s/d 2026-03-31",
		)
		self.assertEqual(sheet["C5"].value, "'=DIST-001")
		self.assertEqual(sheet["D5"].value, "'@DIST-SOURCE")
		self.assertEqual(sheet["E5"].value, "'@Amoxicillin 500 mg")
		self.assertEqual(sheet["F5"].value, "'-Tablet")
		self.assertEqual(sheet["I5"].value, "'+BATCH-LOT")
		self.assertEqual(sheet["L5"].value, "'=BATCH-01")
		self.assertEqual(sheet["A2"].data_type, "s")
		self.assertEqual(sheet["G5"].value, 5)
		self.assertEqual(sheet["H5"].value, 2500)
		self.assertEqual(sheet["K5"].value, 12500)
		self.assertEqual(sheet["G5"].data_type, "n")
		self.assertEqual(sheet["H5"].data_type, "n")
		self.assertEqual(sheet["K5"].data_type, "n")

