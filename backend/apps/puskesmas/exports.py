"""Excel export helpers for the puskesmas reports module."""
from decimal import Decimal

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from apps.core.xlsx_exports import escape_xlsx_formula


# Shared styles (mirrors reports/exports.py style)
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
IDR_FORMAT = "#,##0.00"
NUMBER_FORMAT = "#,##0"


def _cell_value(value):
    return escape_xlsx_formula(value)


def _apply_header_row(ws, row_num, values, col_widths=None):
    """Apply header styling to a row."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=_cell_value(val))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    if col_widths:
        for col_idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w


def _apply_border(ws, row_num, col_count):
    """Apply borders to all cells in a row."""
    for col_idx in range(1, col_count + 1):
        ws.cell(row=row_num, column=col_idx).border = THIN_BORDER


def _make_response(wb, filename):
    """Return an HttpResponse streaming the workbook as .xlsx."""
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_puskesmas_penerimaan_excel(report_data, start_date, end_date, facility_name):
    """Export Riwayat Penerimaan Puskesmas report to Excel.

    Args:
        report_data: list of dicts with keys:
            document_number, received_date, nama_barang, satuan,
            quantity, unit_price, notes
        start_date, end_date: date objects
        facility_name: str — name of the Puskesmas facility
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Riwayat Penerimaan"

    headers = [
        "No",
        "Tanggal SBBK",
        "No. Dokumen",
        "Nama Barang",
        "Satuan",
        "Jumlah Diterima",
        "Harga Satuan (Rp)",
        "Total Nilai (Rp)",
        "Keterangan",
    ]
    col_widths = [6, 18, 24, 32, 10, 16, 20, 22, 28]
    col_count = len(headers)
    last_col = get_column_letter(col_count)

    # Title rows
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws.cell(
        row=1,
        column=1,
        value=_cell_value("RIWAYAT PENERIMAAN BERDASARKAN SBBK PUSKESMAS"),
    )
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A2:{last_col}2")
    period_cell = ws.cell(
        row=2,
        column=1,
        value=_cell_value(f"Fasilitas: {facility_name} | Periode: {start_date} s/d {end_date}"),
    )
    period_cell.font = Font(bold=True, size=11)
    period_cell.alignment = Alignment(horizontal="center")

    _apply_header_row(ws, 4, headers, col_widths)

    row_num = 5
    total_qty = Decimal("0")
    total_value = Decimal("0")

    for idx, row in enumerate(report_data, 1):
        received_date = row.get("received_date")
        received_date_str = received_date.strftime("%d/%m/%Y") if received_date else "-"
        qty = Decimal(str(row.get("quantity", 0) or 0))
        unit_price = Decimal(str(row.get("unit_price", 0) or 0))
        total_price = qty * unit_price

        values = [
            idx,
            received_date_str,
            row.get("document_number", ""),
            row.get("nama_barang", ""),
            row.get("satuan", ""),
            float(qty),
            float(unit_price),
            float(total_price),
            row.get("notes", "") or "-",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=_cell_value(val))
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 6:
                cell.number_format = NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (7, 8):
                cell.number_format = IDR_FORMAT
                cell.alignment = Alignment(horizontal="right")

        total_qty += qty
        total_value += total_price
        row_num += 1

    # Total row
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
    ws.cell(row=row_num, column=2, value=_cell_value("TOTAL")).font = Font(bold=True)
    qty_cell = ws.cell(row=row_num, column=6, value=float(total_qty))
    qty_cell.number_format = NUMBER_FORMAT
    qty_cell.alignment = Alignment(horizontal="right")
    qty_cell.font = Font(bold=True)
    val_cell = ws.cell(row=row_num, column=8, value=float(total_value))
    val_cell.number_format = IDR_FORMAT
    val_cell.alignment = Alignment(horizontal="right")
    val_cell.font = Font(bold=True)

    filename = f"Riwayat_Penerimaan_{facility_name}_{start_date}_{end_date}.xlsx"
    return _make_response(wb, filename)


def export_puskesmas_pemakaian_excel(report_data, year, month_label, facility_name):
    """Export Riwayat Pemakaian Puskesmas report to Excel.

    Args:
        report_data: list of dicts with keys:
            period_display, nama_barang, satuan, stock_awal, penerimaan,
            pemakaian, stock_keseluruhan, permintaan_jumlah
        year: int
        month_label: str — e.g. "Semua Bulan" or "Januari"
        facility_name: str
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Riwayat Pemakaian"

    headers = [
        "No",
        "Periode",
        "Nama Barang",
        "Satuan",
        "Stok Awal",
        "Penerimaan",
        "Pemakaian",
        "Stok Akhir",
        "Permintaan",
    ]
    col_widths = [6, 18, 32, 10, 14, 14, 14, 14, 14]
    col_count = len(headers)
    last_col = get_column_letter(col_count)

    # Title rows
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws.cell(
        row=1,
        column=1,
        value=_cell_value("RIWAYAT PEMAKAIAN PUSKESMAS (DARI DATA LPLPO)"),
    )
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A2:{last_col}2")
    period_cell = ws.cell(
        row=2,
        column=1,
        value=_cell_value(f"Fasilitas: {facility_name} | Tahun: {year} | Bulan: {month_label}"),
    )
    period_cell.font = Font(bold=True, size=11)
    period_cell.alignment = Alignment(horizontal="center")

    _apply_header_row(ws, 4, headers, col_widths)

    row_num = 5
    for idx, row in enumerate(report_data, 1):
        values = [
            idx,
            row.get("period_display", ""),
            row.get("nama_barang", ""),
            row.get("satuan", ""),
            int(row.get("stock_awal", 0) or 0),
            int(row.get("penerimaan", 0) or 0),
            int(row.get("pemakaian", 0) or 0),
            int(row.get("stock_keseluruhan", 0) or 0),
            int(row.get("permintaan_jumlah", 0) or 0),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=_cell_value(val))
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx >= 5:
                cell.number_format = NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="right")
        row_num += 1

    filename = f"Riwayat_Pemakaian_{facility_name}_{year}.xlsx"
    return _make_response(wb, filename)


def export_puskesmas_persediaan_excel(report_data, year, period_label, facility_name):
    """Export Rincian Laporan Persediaan Puskesmas report to Excel.

    Stock data is sourced from the latest LPLPO (stock_keseluruhan) with
    dynamic adjustments for newer distributions. Batch/lot, expiry, funding
    source, and unit price are not tracked in LPLPO and are excluded.

    Args:
        report_data: list of dicts with keys:
            nama_barang, satuan, kategori, stock_keseluruhan
        year: int
        period_label: str — e.g. "Tahunan" or "Triwulan I"
        facility_name: str
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Rincian Persediaan"

    headers = [
        "No",
        "Kategori",
        "Nama Barang",
        "Satuan",
        "Stok Tersedia",
    ]
    col_widths = [6, 24, 36, 10, 16]
    col_count = len(headers)
    last_col = get_column_letter(col_count)

    # Title rows
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws.cell(
        row=1,
        column=1,
        value=_cell_value("RINCIAN LAPORAN PERSEDIAAN OBAT DAN PERBEKALAN KESEHATAN — PUSKESMAS"),
    )
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A2:{last_col}2")
    period_cell = ws.cell(
        row=2,
        column=1,
        value=_cell_value(f"Fasilitas: {facility_name} | Tahun: {year} | Periode: {period_label}"),
    )
    period_cell.font = Font(bold=True, size=11)
    period_cell.alignment = Alignment(horizontal="center")

    _apply_header_row(ws, 4, headers, col_widths)

    row_num = 5
    for idx, row in enumerate(report_data, 1):
        values = [
            idx,
            row.get("kategori", "Lainnya"),
            row.get("nama_barang", ""),
            row.get("satuan", ""),
            int(row.get("stock_keseluruhan", 0) or 0),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=_cell_value(val))
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 5:
                cell.number_format = NUMBER_FORMAT
                cell.alignment = Alignment(horizontal="right")
        row_num += 1

    filename = f"Rincian_Laporan_Persediaan_{facility_name}_{year}_{period_label}.xlsx"
    return _make_response(wb, filename)


def export_puskesmas_rekap_persediaan_excel(rekap_data, totals, year, period_label, facility_name):
    """Export Rekap Laporan Persediaan Puskesmas to Excel.

    Year-to-date summary per category using LPLPO-derived valuation data.

    Args:
        rekap_data: list of dicts with keys:
            kategori, saldo_awal, nilai_terima, nilai_keluar, saldo_akhir
        totals: dict with grand total sums of the valuation fields
        year: int
        period_label: str
        facility_name: str
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Persediaan"

    headers = [
        "No",
        "Uraian",
        f"Saldo Awal {year}\n(Rp)",
        f"Nilai Terima {year}\n(Rp)",
        f"Nilai Keluar {year}\n(Rp)",
        f"Saldo Akhir {year}\n(Rp)",
    ]
    col_widths = [6, 36, 20, 20, 20, 20]
    col_count = len(headers)
    last_col = get_column_letter(col_count)

    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws.cell(
        row=1, column=1,
        value=_cell_value("REKAPITULASI LAPORAN PERSEDIAAN OBAT DAN PERBEKALAN KESEHATAN — PUSKESMAS"),
    )
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A2:{last_col}2")
    period_cell = ws.cell(
        row=2, column=1,
        value=_cell_value(f"Fasilitas: {facility_name} | Tahun: {year} | Periode: {period_label}"),
    )
    period_cell.font = Font(bold=True, size=11)
    period_cell.alignment = Alignment(horizontal="center")

    _apply_header_row(ws, 4, headers, col_widths)
    ws.row_dimensions[4].height = 30

    row_num = 5
    for idx, row in enumerate(rekap_data, 1):
        values = [
            idx,
            row.get("kategori", "Lainnya"),
            float(Decimal(str(row.get("saldo_awal", 0) or 0))),
            float(Decimal(str(row.get("nilai_terima", 0) or 0))),
            float(Decimal(str(row.get("nilai_keluar", 0) or 0))),
            float(Decimal(str(row.get("saldo_akhir", 0) or 0))),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=_cell_value(val))
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx >= 3:
                cell.number_format = IDR_FORMAT
                cell.alignment = Alignment(horizontal="right")
        row_num += 1

    # Totals row
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
    ws.cell(row=row_num, column=2, value=_cell_value("TOTAL")).font = Font(bold=True)
    total_columns = [
        (3, "saldo_awal", IDR_FORMAT, float),
        (4, "nilai_terima", IDR_FORMAT, float),
        (5, "nilai_keluar", IDR_FORMAT, float),
        (6, "saldo_akhir", IDR_FORMAT, float),
    ]
    for col_idx, key, fmt, caster in total_columns:
        c = ws.cell(row=row_num, column=col_idx, value=caster(totals.get(key, 0)))
        c.number_format = fmt
        c.alignment = Alignment(horizontal="right")
        c.font = Font(bold=True)

    filename = f"Rekap_Persediaan_{facility_name}_{year}_{period_label}.xlsx"
    return _make_response(wb, filename)

