"""Excel export helpers for the reports module."""
from decimal import Decimal
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter


# Shared styles
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SD_HEADER_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
CATEGORY_FILL = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
IDR_FORMAT = '#,##0.00'


def _apply_header_row(ws, row_num, values, col_widths=None):
    """Apply header styling to a row."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    if col_widths:
        for col_idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w


def _apply_border(ws, row_num, col_count):
    """Apply borders to all cells in a row."""
    for col_idx in range(1, col_count + 1):
        ws.cell(row=row_num, column=col_idx).border = THIN_BORDER


def _make_response(wb, filename):
    """Return an HttpResponse streaming the workbook as .xlsx."""
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def export_rincian_excel(report_data, start_date, end_date):
    """Export the Rincian (detail) report to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Rincian"

    # Title rows
    ws.merge_cells("A1:L1")
    title_cell = ws.cell(row=1, column=1, value="LAPORAN PERSEDIAAN OBAT DAN PERBEKALAN KESEHATAN (RINCIAN)")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:L2")
    period_cell = ws.cell(row=2, column=1, value=f"Periode: {start_date} s/d {end_date}")
    period_cell.font = Font(bold=True, size=11)
    period_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "No", "Nama Barang", "Satuan", "Batch", "Kedaluwarsa",
        "Sumber Dana", "Harga Satuan", "Stok Awal",
        "Diterima", "Didistribusi", "ED/Rusak", "Stok Akhir"
    ]
    col_widths = [6, 30, 10, 15, 14, 18, 18, 14, 14, 14, 14, 14]
    _apply_header_row(ws, 4, headers, col_widths)

    row_num = 5
    current_category = None
    item_counter = 0

    for row in report_data:
        cat = row.get('item__kategori__name', 'Lainnya')
        if cat != current_category:
            current_category = cat
            item_counter = 0
            # Category row
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=12)
            cat_cell = ws.cell(row=row_num, column=1, value=cat or "KATEGORI TIDAK DIKETAHUI")
            cat_cell.font = Font(bold=True, size=11)
            cat_cell.fill = CATEGORY_FILL
            _apply_border(ws, row_num, 12)
            row_num += 1

        item_counter += 1
        expiry = row.get('expiry_date')
        if expiry and hasattr(expiry, 'strftime'):
            expiry = expiry.strftime("%d/%m/%Y")
        else:
            expiry = "-"

        values = [
            item_counter,
            row.get('item__nama_barang', ''),
            row.get('item__satuan__name', ''),
            row.get('batch_lot', ''),
            expiry,
            row.get('sumber_dana__name', ''),
            float(row.get('unit_price', 0)),
            float(row.get('initial_stock', 0)),
            float(row.get('received', 0)),
            float(row.get('distributed', 0)),
            float(row.get('expired', 0)),
            float(row.get('ending_stock', 0)),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx >= 7:  # Numeric columns
                cell.number_format = IDR_FORMAT
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
        row_num += 1

    filename = f"Laporan_Rincian_{start_date}_{end_date}.xlsx"
    return _make_response(wb, filename)


def export_rekap_excel(rekap_data, grand_totals, start_date, end_date):
    """Export the Rekap (summary) report to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Rekap"

    # Title rows
    ws.merge_cells("A1:G1")
    title_cell = ws.cell(row=1, column=1, value="LAPORAN PERSEDIAAN OBAT DAN PERBEKALAN KESEHATAN (REKAP)")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    period_cell = ws.cell(row=2, column=1, value=f"Periode: {start_date} s/d {end_date}")
    period_cell.font = Font(bold=True, size=11)
    period_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "No", "Uraian", "Saldo Awal (Rp)", "Nilai Terima (Rp)",
        "Nilai Distribusi (Rp)", "Nilai ED/Rusak (Rp)", "Saldo Akhir (Rp)"
    ]
    col_widths = [6, 28, 24, 24, 24, 24, 24]
    _apply_header_row(ws, 4, headers, col_widths)

    row_num = 5
    for sd_group in rekap_data:
        # Sumber Dana subtotal header
        sd_values = [
            "",
            sd_group['sd_name'],
            float(sd_group['subtotal_saldo_awal']),
            float(sd_group['subtotal_nilai_terima']),
            float(sd_group['subtotal_nilai_distribusi']),
            float(sd_group['subtotal_nilai_ed']),
            float(sd_group['subtotal_saldo_akhir']),
        ]
        for col_idx, val in enumerate(sd_values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = Font(bold=True)
            cell.fill = SD_HEADER_FILL
            cell.border = THIN_BORDER
            if col_idx >= 3:
                cell.number_format = IDR_FORMAT
                cell.alignment = Alignment(horizontal="right")
        row_num += 1

        # Category rows
        for idx, cat in enumerate(sd_group['categories'], 1):
            cat_values = [
                idx,
                cat['kategori'],
                float(cat['saldo_awal']),
                float(cat['nilai_terima']),
                float(cat['nilai_distribusi']),
                float(cat['nilai_ed']),
                float(cat['saldo_akhir']),
            ]
            for col_idx, val in enumerate(cat_values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = THIN_BORDER
                if col_idx >= 3:
                    cell.number_format = IDR_FORMAT
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx == 1:
                    cell.alignment = Alignment(horizontal="center")
            row_num += 1

    # Grand total row
    total_values = [
        "",
        "Total",
        float(grand_totals.get('saldo_awal', 0)),
        float(grand_totals.get('nilai_terima', 0)),
        float(grand_totals.get('nilai_distribusi', 0)),
        float(grand_totals.get('nilai_ed', 0)),
        float(grand_totals.get('saldo_akhir', 0)),
    ]
    for col_idx, val in enumerate(total_values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        if col_idx >= 3:
            cell.number_format = IDR_FORMAT
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 2:
            cell.alignment = Alignment(horizontal="center")

    filename = f"Laporan_Rekap_{start_date}_{end_date}.xlsx"
    return _make_response(wb, filename)
