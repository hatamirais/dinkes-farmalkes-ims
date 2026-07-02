from io import BytesIO
from datetime import date, datetime
from decimal import Decimal

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.distribution.models import Distribution
from apps.items.models import Category, Facility, FundingSource, Item, Location, Supplier, Unit
from apps.procurement.models import ProcurementContract
from apps.receiving.models import Receiving, ReceivingItem
from apps.reports.exports import export_numbering_history_excel, export_pengeluaran_excel
from apps.stock.models import Stock
from apps.users.models import User


class NumberingHistoryReportTests(TestCase):
	@classmethod
	def setUpTestData(cls):
		cls.user = User.objects.create_superuser(
			username="reports-admin",
			password="secret12345",
		)
		cls.unit = Unit.objects.create(code="TAB", name="Tablet")
		cls.category = Category.objects.create(
			code="REPORT-CAT", name="Report Category", sort_order=1
		)
		cls.item = Item.objects.create(
			nama_barang="Paracetamol 500mg",
			satuan=cls.unit,
			kategori=cls.category,
		)
		cls.location = Location.objects.create(code="REP-LOC", name="Gudang Laporan")
		cls.funding_source = FundingSource.objects.create(code="BOK", name="BOK")
		cls.facility = Facility.objects.create(code="PKM-REP", name="Puskesmas Laporan")
		cls.stock = Stock.objects.create(
			item=cls.item,
			location=cls.location,
			batch_lot="REP-01",
			expiry_date="2027-12-31",
			quantity=10,
			reserved=0,
			unit_price=1000,
			sumber_dana=cls.funding_source,
		)

	def setUp(self):
		self.client.force_login(self.user)

	def _create_distribution(self, distribution_type, document_number=None):
		dist = Distribution.objects.create(
			distribution_type=distribution_type,
			document_number=document_number or "",
			request_date="2026-04-01",
			facility=self.facility,
			status=Distribution.Status.DRAFT,
			created_by=self.user,
			notes="Catatan ringkas",
		)
		dist.items.create(
			item=self.item,
			quantity_requested=5,
			quantity_approved=5,
			stock=self.stock,
		)
		return dist

	def test_numbering_history_page_lists_lplpo_and_special_request(self):
		lplpo_dist = self._create_distribution(Distribution.DistributionType.LPLPO)
		special_dist = self._create_distribution(Distribution.DistributionType.SPECIAL_REQUEST)

		response = self.client.get(reverse('reports:numbering_history'))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, lplpo_dist.document_number)
		self.assertContains(response, special_dist.document_number)
		self.assertContains(response, "Riwayat Penomoran")
		self.assertContains(response, "Lihat Dokumen")

	def test_numbering_history_page_filters_by_document_type(self):
		lplpo_dist = self._create_distribution(Distribution.DistributionType.LPLPO)
		self._create_distribution(Distribution.DistributionType.SPECIAL_REQUEST)

		response = self.client.get(
			reverse('reports:numbering_history'),
			{'distribution_type': Distribution.DistributionType.LPLPO, 'year': 2026},
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, lplpo_dist.document_number)
		self.assertNotContains(response, 'KD.F/2026')

	def test_numbering_history_page_shows_print_and_export_actions(self):
		self._create_distribution(Distribution.DistributionType.LPLPO)

		response = self.client.get(reverse('reports:numbering_history'))

		self.assertContains(response, 'Cetak Laporan')
		self.assertContains(response, 'Export Excel')

	def test_numbering_history_excel_export_returns_workbook(self):
		self._create_distribution(Distribution.DistributionType.LPLPO)

		response = self.client.get(
			reverse('reports:numbering_history'),
			{'year': 2026, 'format': 'excel'},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(
			response['Content-Type'],
			'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
		)
		self.assertIn('Riwayat_Penomoran_2026.xlsx', response['Content-Disposition'])

	def test_numbering_history_excel_neutralizes_formula_prefixed_strings(self):
		response = export_numbering_history_excel(
			[
				{
					"document_number": "=DOC-001",
					"distribution_type": "+LPLPO",
					"status": "@Draft",
					"facility_name": "-Facility",
					"source_label": "=LPLPO",
					"source_document_number": "=SRC-001",
					"created_at": None,
					"item_count": 1,
				}
			],
			2026,
			"=Semua Dokumen",
		)

		workbook = load_workbook(BytesIO(response.content))
		sheet = workbook.active

		self.assertEqual(sheet["A2"].value, "Tahun: 2026 | Jenis Dokumen: =Semua Dokumen")
		self.assertEqual(sheet["B5"].value, "'=DOC-001")
		self.assertEqual(sheet["C5"].value, "'+LPLPO")
		self.assertEqual(sheet["D5"].value, "'@Draft")
		self.assertEqual(sheet["E5"].value, "'-Facility")
		self.assertEqual(sheet["F5"].value, "'=LPLPO: =SRC-001")
		self.assertEqual(sheet["A2"].data_type, "s")
		self.assertEqual(sheet["B5"].data_type, "s")


class PengeluaranReportTests(TestCase):
	@classmethod
	def setUpTestData(cls):
		cls.user = User.objects.create_superuser(
			username="pengeluaran-admin",
			password="secret12345",
		)
		cls.unit = Unit.objects.create(code="BOT", name="Botol")
		cls.category = Category.objects.create(
			code="OUT-CAT", name="Outbound Category", sort_order=1
		)
		cls.item = Item.objects.create(
			nama_barang="Amoxicillin Syrup",
			satuan=cls.unit,
			kategori=cls.category,
		)
		cls.location = Location.objects.create(code="OUT-LOC", name="Gudang Pengeluaran")
		cls.funding_source = FundingSource.objects.create(code="DAU", name="DAU")
		cls.facility = Facility.objects.create(code="PKM-OUT", name="Puskesmas Pengeluaran")
		cls.other_facility = Facility.objects.create(code="PKM-ALT", name="Puskesmas Alternatif")
		cls.stock = Stock.objects.create(
			item=cls.item,
			location=cls.location,
			batch_lot="OUT-01",
			expiry_date="2027-10-31",
			quantity=50,
			reserved=0,
			unit_price=2500,
			sumber_dana=cls.funding_source,
		)

	def setUp(self):
		self.client.force_login(self.user)

	def _create_distribution(self, distribution_type, facility=None, document_number=None):
		dist = Distribution.objects.create(
			distribution_type=distribution_type,
			document_number=document_number or "",
			request_date="2026-04-15",
			facility=facility or self.facility,
			status=Distribution.Status.DISTRIBUTED,
			created_by=self.user,
			notes="Pengeluaran terverifikasi",
		)
		dist.items.create(
			item=self.item,
			quantity_requested=7,
			quantity_approved=5,
			stock=self.stock,
		)
		return dist

	def test_pengeluaran_report_filters_by_distribution_type(self):
		allocation_dist = self._create_distribution(
			Distribution.DistributionType.ALLOCATION,
			document_number="ALLOC-REP-001",
		)
		self._create_distribution(
			Distribution.DistributionType.LPLPO,
			document_number="LPLPO-REP-001",
		)

		response = self.client.get(
			reverse('reports:pengeluaran'),
			{
				'start_date': '2026-04-01',
				'end_date': '2026-04-30',
				'distribution_type': Distribution.DistributionType.ALLOCATION,
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, allocation_dist.document_number)
		self.assertNotContains(response, 'LPLPO-REP-001')

	def test_pengeluaran_report_combined_view_remains_available(self):
		allocation_dist = self._create_distribution(
			Distribution.DistributionType.ALLOCATION,
			document_number="ALLOC-REP-ALL",
		)
		lplpo_dist = self._create_distribution(
			Distribution.DistributionType.LPLPO,
			document_number="LPLPO-REP-ALL",
		)

		response = self.client.get(
			reverse('reports:pengeluaran'),
			{
				'start_date': '2026-04-01',
				'end_date': '2026-04-30',
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, allocation_dist.document_number)
		self.assertContains(response, lplpo_dist.document_number)
		self.assertContains(response, 'Semua Distribusi')
		self.assertContains(response, 'Permintaan Khusus')
		self.assertContains(response, 'Alokasi')
		self.assertContains(response, 'LPLPO')

	def test_pengeluaran_report_combines_facility_and_distribution_type_filters(self):
		matching_dist = self._create_distribution(
			Distribution.DistributionType.SPECIAL_REQUEST,
			facility=self.facility,
			document_number="SPEC-REP-001",
		)
		self._create_distribution(
			Distribution.DistributionType.SPECIAL_REQUEST,
			facility=self.other_facility,
			document_number="SPEC-REP-ALT",
		)
		self._create_distribution(
			Distribution.DistributionType.ALLOCATION,
			facility=self.facility,
			document_number="ALLOC-REP-FAC",
		)

		response = self.client.get(
			reverse('reports:pengeluaran'),
			{
				'start_date': '2026-04-01',
				'end_date': '2026-04-30',
				'facility': self.facility.pk,
				'distribution_type': Distribution.DistributionType.SPECIAL_REQUEST,
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, matching_dist.document_number)
		self.assertNotContains(response, 'SPEC-REP-ALT')
		self.assertNotContains(response, 'ALLOC-REP-FAC')

	def test_pengeluaran_report_invalid_distribution_type_keeps_report_empty(self):
		self._create_distribution(
			Distribution.DistributionType.ALLOCATION,
			document_number="ALLOC-REP-INVALID",
		)

		response = self.client.get(
			reverse('reports:pengeluaran'),
			{
				'start_date': '2026-04-01',
				'end_date': '2026-04-30',
				'distribution_type': 'NOT_A_REAL_TYPE',
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertFalse(response.context['form'].is_valid())
		self.assertEqual(response.context['report_data'], [])
		self.assertNotContains(response, 'ALLOC-REP-INVALID')

	def test_pengeluaran_report_excel_export_uses_active_tab_label(self):
		self._create_distribution(
			Distribution.DistributionType.ALLOCATION,
			document_number="ALLOC-REP-EXPORT",
		)

		response = self.client.get(
			reverse('reports:pengeluaran'),
			{
				'start_date': '2026-04-01',
				'end_date': '2026-04-30',
				'distribution_type': Distribution.DistributionType.ALLOCATION,
				'format': 'excel',
			},
		)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(
			response['Content-Type'],
			'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
		)
		self.assertIn(
			'Laporan_Pengeluaran_Alokasi_2026-04-01_2026-04-30.xlsx',
			response['Content-Disposition'],
		)

	def test_pengeluaran_report_tabs_include_distribution_type_query(self):
		response = self.client.get(
			reverse('reports:pengeluaran'),
			{
				'start_date': '2026-04-01',
				'end_date': '2026-04-30',
			},
		)

		self.assertEqual(response.status_code, 200)
		tabs = response.context['tabs']
		self.assertTrue(any(tab['value'] == '' for tab in tabs))
		self.assertTrue(any(tab['value'] == Distribution.DistributionType.ALLOCATION for tab in tabs))

		allocation_tab = next(
			tab for tab in tabs if tab['value'] == Distribution.DistributionType.ALLOCATION
		)
		self.assertIn('distribution_type=ALLOCATION', allocation_tab['url'])
		self.assertIn('start_date=2026-04-01', allocation_tab['url'])
		self.assertIn('end_date=2026-04-30', allocation_tab['url'])

	def test_pengeluaran_excel_neutralizes_formula_prefixed_text_and_keeps_numeric_cells(self):
		response = export_pengeluaran_excel(
			[
				{
					"document_number": "=DOC-OUT-1",
					"facility_name": "+Puskesmas Formula",
					"nama_barang": "@Amoxicillin",
					"satuan": "-Botol",
					"batch_lot": "=BATCH-01",
					"expiry_date": None,
					"sumber_dana": "=DAU",
					"unit_price": 2500,
					"quantity": 5,
					"total_price": 12500,
				}
			],
			"2026-04-01",
			"2026-04-30",
			facility_name="=Semua Fasilitas",
			distribution_type_label="+Semua Distribusi",
		)

		workbook = load_workbook(BytesIO(response.content))
		sheet = workbook.active

		self.assertEqual(
			sheet["A2"].value,
			"Periode: 2026-04-01 s/d 2026-04-30 | Fasilitas: =Semua Fasilitas | Jenis Distribusi: +Semua Distribusi",
		)
		self.assertEqual(sheet["B5"].value, "'=DOC-OUT-1")
		self.assertEqual(sheet["C5"].value, "'+Puskesmas Formula")
		self.assertEqual(sheet["D5"].value, "'@Amoxicillin")
		self.assertEqual(sheet["E5"].value, "'-Botol")
		self.assertEqual(sheet["F5"].value, "'=BATCH-01")
		self.assertEqual(sheet["H5"].value, "'=DAU")
		self.assertEqual(sheet["A2"].data_type, "s")
		self.assertEqual(sheet["I5"].value, 2500)
		self.assertEqual(sheet["J5"].value, 5)
		self.assertEqual(sheet["K5"].value, 12500)
		self.assertEqual(sheet["I5"].data_type, "n")
		self.assertEqual(sheet["J5"].data_type, "n")
		self.assertEqual(sheet["K5"].data_type, "n")


class ProcurementReportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_superuser(
            username="proc-report-admin",
            password="secret12345",
        )
        cls.unit = Unit.objects.create(code="PRT", name="Pcs")
        cls.category = Category.objects.create(code="PRC", name="Procurement Category", sort_order=3)
        cls.item = Item.objects.create(
            kode_barang="REP-PROC-001",
            nama_barang="Vitamin C 500mg",
            satuan=cls.unit,
            kategori=cls.category,
        )
        cls.location = Location.objects.create(code="REP-PROC-LOC", name="Gudang Pengadaan")
        cls.funding = FundingSource.objects.create(code="APBD", name="APBD")
        cls.supplier = Supplier.objects.create(code="REP-SUP", name="PT Report Supplier")
        cls.contract = ProcurementContract.objects.create(
            document_number="SPJ-2026-00077",
            contract_date="2026-07-01",
            supplier=cls.supplier,
            sumber_dana=cls.funding,
            status=ProcurementContract.Status.APPROVED,
            created_by=cls.user,
            approved_by=cls.user,
        )
        cls.receiving = Receiving.objects.create(
            document_number="RCV-2026-00999",
            receiving_type=Receiving.ReceivingType.PROCUREMENT,
            receiving_date="2026-07-10",
            supplier=cls.supplier,
            sumber_dana=cls.funding,
            status=Receiving.Status.VERIFIED,
            is_planned=False,
            contract=cls.contract,
            created_by=cls.user,
            verified_by=cls.user,
        )
        ReceivingItem.objects.create(
            receiving=cls.receiving,
            item=cls.item,
            quantity=5,
            batch_lot="REP-BATCH-001",
            expiry_date="2030-01-01",
            unit_price=1200,
            location=cls.location,
            received_by=cls.user,
        )

    def setUp(self):
        self.client.force_login(self.user)

    def test_procurement_report_shows_contract_reference(self):
        response = self.client.get(
            reverse("reports:pengadaan"),
            {"start_date": "2026-07-01", "end_date": "2026-07-31"},
            secure=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "SPJ-2026-00077")
        self.assertEqual(response.context["report_data"][0]["contract_document_number"], "SPJ-2026-00077")

    def test_procurement_report_excel_includes_spj_column(self):
        response = self.client.get(
            reverse("reports:pengadaan"),
            {"start_date": "2026-07-01", "end_date": "2026-07-31", "format": "excel"},
            secure=True,
        )

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        self.assertEqual(sheet["C4"].value, "No. SPJ")
        self.assertEqual(sheet["C5"].value, "SPJ-2026-00077")


class ProcurementReceivingReportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_superuser(
            username="report-proc-admin",
            password="secret12345",
        )
        cls.unit = Unit.objects.create(code="PRT", name="Pcs")
        cls.category = Category.objects.create(
            code="PROC-REP",
            name="Procurement Report",
            sort_order=2,
        )
        cls.item = Item.objects.create(
            kode_barang="PROC-REP-001",
            nama_barang="Rapid Test",
            satuan=cls.unit,
            kategori=cls.category,
            minimum_stock=0,
        )
        cls.location = Location.objects.create(code="PROC-REP-LOC", name="Gudang Report")
        cls.funding_source = FundingSource.objects.create(code="PROC-REP-FS", name="Procurement FS")
        cls.supplier = Supplier.objects.create(code="PROC-REP-SUP", name="Supplier Report")

    def setUp(self):
        self.client.force_login(self.user)

    def test_pengadaan_report_includes_contract_reference(self):
        from apps.procurement.models import ProcurementContract, ProcurementContractLine
        from apps.procurement.services import approve_contract
        from apps.receiving.models import Receiving, ReceivingOrderItem

        contract = ProcurementContract.objects.create(
            document_number="",
            contract_date=date(2026, 6, 28),
            supplier=self.supplier,
            sumber_dana=self.funding_source,
            notes="Kontrak report",
            created_by=self.user,
            status=ProcurementContract.Status.SUBMITTED,
            submitted_by=self.user,
            submitted_at=timezone.now(),
        )
        line = ProcurementContractLine.objects.create(
            contract=contract,
            item=self.item,
            original_quantity=Decimal("8"),
            original_unit_price=Decimal("11000"),
        )
        approve_contract(contract, self.user)
        receiving = Receiving.objects.get(contract=contract)
        order_item = ReceivingOrderItem.objects.get(receiving=receiving, contract_line=line)
        receiving.receiving_date = date(2026, 6, 28)
        receiving.save(update_fields=["receiving_date", "updated_at"])

        response = self.client.post(
            reverse("receiving:receiving_plan_receive", args=[receiving.pk]),
            {
                "items-TOTAL_FORMS": "1",
                "items-INITIAL_FORMS": "0",
                "items-MIN_NUM_FORMS": "0",
                "items-MAX_NUM_FORMS": "1000",
                "items-0-order_item": str(order_item.pk),
                "items-0-quantity": "8",
                "items-0-batch_lot": "PROC-REPORT-BATCH",
                "items-0-expiry_date": "2031-12-31",
                "items-0-unit_price": "11000",
                "items-0-location": str(self.location.pk),
            },
            secure=True,
        )
        self.assertEqual(response.status_code, 302)
        receipt = ReceivingItem.objects.get(receiving=receiving)
        receipt.received_at = timezone.make_aware(datetime(2026, 7, 10, 9, 0, 0))
        receipt.save(update_fields=["received_at"])

        page = self.client.get(
            reverse("reports:pengadaan"),
            {"start_date": "2026-07-01", "end_date": "2026-07-31"},
            secure=True,
        )

        self.assertEqual(page.status_code, 200)
        self.assertContains(page, contract.document_number)
        self.assertContains(page, receiving.document_number)
        self.assertContains(page, "NO. SPJ")
        self.assertEqual(page.context["report_data"][0]["receiving_date"], date(2026, 7, 10))
