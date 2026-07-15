from io import BytesIO

from django.contrib.admin.sites import AdminSite
from django.core.exceptions import ValidationError
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook
from tablib import Dataset

from apps.core.csv_exports import SanitizedCSV
from apps.items.admin import ItemAdmin, ItemResource
from apps.items.forms import ItemForm
from apps.items.models import Category, Item, Program, TherapeuticClass, Unit
from apps.users.models import ModuleAccess, User


class ItemAdminCsvExportSecurityTest(TestCase):
    def setUp(self):
        self.unit = Unit.objects.create(code="TAB", name="Tablet")
        self.category = Category.objects.create(code="OBAT", name="Obat")

    def test_item_admin_uses_sanitized_csv_format(self):
        admin = ItemAdmin(Item, AdminSite())

        self.assertIn(SanitizedCSV, admin.get_export_formats())

    def test_item_resource_csv_export_neutralizes_formula_prefixed_values(self):
        item = Item.objects.create(
            nama_barang='=HYPERLINK("http://example.com")',
            satuan=self.unit,
            kategori=self.category,
        )

        dataset = ItemResource().export(Item.objects.filter(pk=item.pk))
        csv_output = SanitizedCSV().export_data(dataset)

        self.assertIn("\"'=HYPERLINK", csv_output)
        self.assertIn(item.kode_barang, csv_output)


class LookupModelGuardrailTest(TestCase):
    def test_unit_save_normalizes_code_and_name(self):
        unit = Unit.objects.create(code=" tab ", name=" Tablet   Oral ")
        self.assertEqual(unit.code, "TAB")
        self.assertEqual(unit.name, "Tablet Oral")

    def test_category_duplicate_name_case_insensitive_blocked(self):
        Category.objects.create(code="TABLET", name="Tablet")
        duplicate = Category(code="TABLET2", name="tablet")
        with self.assertRaises(ValidationError):
            duplicate.save()

    def test_program_duplicate_name_case_insensitive_blocked(self):
        Program.objects.create(code="HIV", name="Human Immunodeficiency Virus")
        duplicate = Program(code="HIV2", name="human immunodeficiency virus")
        with self.assertRaises(ValidationError):
            duplicate.save()

    def test_therapeutic_class_duplicate_name_case_insensitive_blocked(self):
        TherapeuticClass.objects.create(code="ABX", name="Antibiotik")
        duplicate = TherapeuticClass(code="ABX2", name="antibiotik")
        with self.assertRaises(ValidationError):
            duplicate.save()


class ItemTherapeuticClassTests(TestCase):
    def setUp(self):
        self.unit = Unit.objects.create(code="TAB", name="Tablet")
        self.category = Category.objects.create(code="OBAT", name="Obat")
        self.therapy_antibiotic = TherapeuticClass.objects.create(
            code="ABX",
            name="Antibiotik",
        )
        self.therapy_respiratory = TherapeuticClass.objects.create(
            code="RESP",
            name="Respirasi",
        )

    def test_item_form_exposes_therapeutic_classes_field(self):
        form = ItemForm()
        self.assertIn("barcode", form.fields)
        self.assertIn("therapeutic_classes", form.fields)
        self.assertIn("requires_expiry_date", form.fields)
        self.assertTrue(form.fields["requires_expiry_date"].initial)

    def test_item_form_accepts_blank_barcode_as_null(self):
        form = ItemForm(
            data={
                "barcode": "",
                "nama_barang": "Paracetamol",
                "satuan": self.unit.pk,
                "kategori": self.category.pk,
                "minimum_stock": "0",
                "description": "",
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        item = form.save()
        self.assertIsNone(item.barcode)

    def test_item_form_saves_barcode(self):
        form = ItemForm(
            data={
                "barcode": "8991234567890",
                "nama_barang": "Paracetamol",
                "satuan": self.unit.pk,
                "kategori": self.category.pk,
                "minimum_stock": "0",
                "description": "",
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        item = form.save()
        self.assertEqual(item.barcode, "8991234567890")

    def test_item_form_rejects_duplicate_non_null_barcode(self):
        Item.objects.create(
            nama_barang="Amoxicillin",
            satuan=self.unit,
            kategori=self.category,
            barcode="8991234567890",
        )
        form = ItemForm(
            data={
                "barcode": "8991234567890",
                "nama_barang": "Paracetamol",
                "satuan": self.unit.pk,
                "kategori": self.category.pk,
                "minimum_stock": "0",
                "description": "",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("barcode", form.errors)

    def test_item_can_store_multiple_therapeutic_classes(self):
        item = Item.objects.create(
            nama_barang="Amoxicillin",
            satuan=self.unit,
            kategori=self.category,
        )
        item.therapeutic_classes.set(
            [self.therapy_antibiotic, self.therapy_respiratory]
        )

        self.assertQuerySetEqual(
            item.therapeutic_classes.order_by("code").values_list("code", flat=True),
            ["ABX", "RESP"],
            transform=lambda value: value,
        )

    def test_item_resource_exports_pipe_separated_therapeutic_codes(self):
        item = Item.objects.create(
            nama_barang="Amoxicillin",
            satuan=self.unit,
            kategori=self.category,
            barcode="8991234567890",
        )
        item.therapeutic_classes.set(
            [self.therapy_antibiotic, self.therapy_respiratory]
        )

        dataset = ItemResource().export(Item.objects.filter(pk=item.pk))

        self.assertEqual(dataset.dict[0]["barcode"], "8991234567890")
        self.assertEqual(dataset.dict[0]["therapeutic_classes"], "ABX|RESP")

    def test_item_resource_imports_multiple_therapeutic_codes(self):
        csv = (
            "barcode,nama_barang,satuan,kategori,is_program_item,program,therapeutic_classes,minimum_stock,description,is_active\n"
            "8991234567890,Paracetamol,TAB,OBAT,0,,ABX|RESP,0,desc,1\n"
        )
        dataset = Dataset().load(csv, format="csv")

        result = ItemResource().import_data(dataset, dry_run=False, raise_errors=True)

        self.assertFalse(result.has_errors())
        item = Item.objects.get(nama_barang="Paracetamol")
        self.assertEqual(item.barcode, "8991234567890")
        self.assertQuerySetEqual(
            item.therapeutic_classes.order_by("code").values_list("code", flat=True),
            ["ABX", "RESP"],
            transform=lambda value: value,
        )

    def test_item_resource_import_without_barcode_column_preserves_existing_barcode(self):
        item = Item.objects.create(
            nama_barang="Paracetamol",
            satuan=self.unit,
            kategori=self.category,
            barcode="8991234567890",
            minimum_stock=1,
        )
        csv = (
            "nama_barang,satuan,kategori,is_program_item,program,therapeutic_classes,minimum_stock,description,is_active\n"
            "Paracetamol,TAB,OBAT,0,,ABX,5,updated,1\n"
        )
        dataset = Dataset().load(csv, format="csv")

        result = ItemResource().import_data(dataset, dry_run=False, raise_errors=True)

        self.assertFalse(result.has_errors())
        item.refresh_from_db()
        self.assertEqual(item.barcode, "8991234567890")
        self.assertEqual(item.minimum_stock, 5)

    def test_item_resource_rejects_unknown_therapeutic_code(self):
        csv = (
            "nama_barang,satuan,kategori,is_program_item,program,therapeutic_classes,minimum_stock,description,is_active\n"
            "Paracetamol,TAB,OBAT,0,,UNKNOWN,0,desc,1\n"
        )
        dataset = Dataset().load(csv, format="csv")

        with self.assertRaises(Exception) as exc:
            ItemResource().import_data(dataset, dry_run=False, raise_errors=True)

        self.assertIn("Kode kelas terapi tidak ditemukan", str(exc.exception))


class ItemEssentialTagTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(
            username="admin",
            email="admin@example.com",
            password="password12345",
        )
        self.client.force_login(self.user)
        self.unit = Unit.objects.create(code="TAB", name="Tablet")
        self.category = Category.objects.create(code="OBAT", name="Obat")
        self.therapeutic_class = TherapeuticClass.objects.create(
            code="ANLG",
            name="Analgesik",
        )

    def test_item_defaults_to_not_essential(self):
        item = Item.objects.create(
            nama_barang="Paracetamol",
            satuan=self.unit,
            kategori=self.category,
        )
        self.assertFalse(item.is_essential)

    def test_item_list_shows_essential_badge_and_therapeutic_class(self):
        item = Item.objects.create(
            nama_barang="Amoxicillin",
            satuan=self.unit,
            kategori=self.category,
            is_essential=True,
            barcode="8991234567890",
        )
        item.therapeutic_classes.add(self.therapeutic_class)

        response = self.client.get(reverse("items:item_list"), secure=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "8991234567890")
        self.assertContains(response, "[E] Esensial")
        self.assertContains(response, "Analgesik")

    def test_item_list_search_matches_barcode(self):
        matching = Item.objects.create(
            nama_barang="Amoxicillin",
            satuan=self.unit,
            kategori=self.category,
            barcode="8991234567890",
        )
        other = Item.objects.create(
            nama_barang="Vitamin C",
            satuan=self.unit,
            kategori=self.category,
            barcode="8999999999999",
        )

        response = self.client.get(
            reverse("items:item_list"),
            {"q": "8991234567890"},
            secure=True,
        )

        self.assertContains(response, matching.nama_barang)
        self.assertNotContains(response, other.nama_barang)

    def test_item_list_filters_by_therapeutic_class(self):
        item = Item.objects.create(
            nama_barang="Amoxicillin",
            satuan=self.unit,
            kategori=self.category,
        )
        item.therapeutic_classes.add(self.therapeutic_class)
        other = Item.objects.create(
            nama_barang="Vitamin C",
            satuan=self.unit,
            kategori=self.category,
        )

        response = self.client.get(
            reverse("items:item_list"),
            {"therapeutic_class": self.therapeutic_class.pk},
            secure=True,
        )

        self.assertContains(response, item.nama_barang)
        self.assertNotContains(response, other.nama_barang)

    def test_item_list_filters_essential_items(self):
        essential_item = Item.objects.create(
            nama_barang="Amoxicillin",
            satuan=self.unit,
            kategori=self.category,
            is_essential=True,
        )
        non_essential_item = Item.objects.create(
            nama_barang="Vitamin C",
            satuan=self.unit,
            kategori=self.category,
            is_essential=False,
        )

        response = self.client.get(
            reverse("items:item_list"),
            {"essential": "1"},
            secure=True,
        )

        self.assertContains(response, essential_item.nama_barang)
        self.assertNotContains(response, non_essential_item.nama_barang)

    def test_item_list_filters_non_essential_items(self):
        essential_item = Item.objects.create(
            nama_barang="Amoxicillin",
            satuan=self.unit,
            kategori=self.category,
            is_essential=True,
        )
        non_essential_item = Item.objects.create(
            nama_barang="Vitamin C",
            satuan=self.unit,
            kategori=self.category,
            is_essential=False,
        )

        response = self.client.get(
            reverse("items:item_list"),
            {"essential": "0"},
            secure=True,
        )

        self.assertContains(response, non_essential_item.nama_barang)
        self.assertNotContains(response, essential_item.nama_barang)

    def test_item_list_without_essential_filter_shows_both_types(self):
        essential_item = Item.objects.create(
            nama_barang="Amoxicillin",
            satuan=self.unit,
            kategori=self.category,
            is_essential=True,
        )
        non_essential_item = Item.objects.create(
            nama_barang="Vitamin C",
            satuan=self.unit,
            kategori=self.category,
            is_essential=False,
        )

        response = self.client.get(reverse("items:item_list"), secure=True)

        self.assertContains(response, essential_item.nama_barang)
        self.assertContains(response, non_essential_item.nama_barang)

    def test_item_export_requires_login(self):
        self.client.logout()

        response = self.client.get(reverse("items:item_export"), secure=True)

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("login"), response["Location"])

    def test_item_export_requires_permission(self):
        self.client.logout()
        user = User.objects.create_user(
            username="viewer",
            email="viewer@example.com",
            password="password12345",
            role=User.Role.PUSKESMAS,
        )
        ModuleAccess.objects.update_or_create(
            user=user,
            module=ModuleAccess.Module.ITEMS,
            defaults={"scope": ModuleAccess.Scope.NONE},
        )
        self.client.force_login(user)

        response = self.client.get(reverse("items:item_export"), secure=True)

        self.assertEqual(response.status_code, 403)

    def test_item_export_returns_xlsx_with_essential_and_therapeutic_columns(self):
        item = Item.objects.create(
            nama_barang="Amoxicillin",
            satuan=self.unit,
            kategori=self.category,
            is_essential=True,
            barcode="8991234567890",
        )
        item.therapeutic_classes.add(self.therapeutic_class)

        response = self.client.get(reverse("items:item_export"), secure=True)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn('attachment; filename="Daftar_Barang.xlsx"', response["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active

        self.assertEqual(sheet.title, "Daftar Barang")
        self.assertEqual(sheet["B3"].value, "Barcode")
        self.assertEqual(sheet["G3"].value, "Esensial")
        self.assertEqual(sheet["H3"].value, "Terapi Obat")
        self.assertEqual(sheet["A4"].value, item.kode_barang)
        self.assertEqual(sheet["B4"].value, item.barcode)
        self.assertEqual(sheet["C4"].value, item.nama_barang)
        self.assertEqual(sheet["G4"].value, "Ya")
        self.assertEqual(sheet["H4"].value, "Analgesik")

    def test_item_export_respects_essential_filter(self):
        essential_item = Item.objects.create(
            nama_barang="Amoxicillin",
            satuan=self.unit,
            kategori=self.category,
            is_essential=True,
        )
        Item.objects.create(
            nama_barang="Vitamin C",
            satuan=self.unit,
            kategori=self.category,
            is_essential=False,
        )

        response = self.client.get(
            reverse("items:item_export"),
            {"essential": "1"},
            secure=True,
        )

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        exported_names = [cell for cell in (sheet["C4"].value, sheet["C5"].value) if cell]

        self.assertEqual(exported_names, [essential_item.nama_barang])

    def test_item_export_sanitizes_formula_prefixed_values(self):
        item = Item.objects.create(
            nama_barang="=SUM(1,1)",
            satuan=self.unit,
            kategori=self.category,
        )

        response = self.client.get(reverse("items:item_export"), secure=True)

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active

        self.assertEqual(sheet["C4"].value, "'=SUM(1,1)")
        self.assertEqual(sheet["A4"].value, item.kode_barang)

    def test_picker_label_strips_program_and_essential_suffixes(self):
        item = Item.objects.create(
            nama_barang="Isoniazid (H) 300 mg Tablet [P] [E]",
            satuan=self.unit,
            kategori=self.category,
        )

        self.assertEqual(item.picker_label, "Isoniazid (H) 300 mg Tablet")


class TherapeuticClassQuickCreateTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(
            username="admin",
            email="admin@example.com",
            password="password12345",
        )
        self.client.force_login(self.user)

    @override_settings(ITEM_MUTATION_RATE_LIMIT="1/m", RATELIMIT_USE_CACHE="locmem")
    def test_quick_create_therapeutic_class_uses_item_mutation_rate_limit(self):
        first = self.client.post(
            reverse("items:quick_create_therapeutic_class"),
            {
                "code": "ABX",
                "name": "Antibiotik",
                "description": "Terapi infeksi",
            },
            secure=True,
        )
        second = self.client.post(
            reverse("items:quick_create_therapeutic_class"),
            {
                "code": "RESP",
                "name": "Respirasi",
                "description": "Terapi saluran napas",
            },
            secure=True,
        )

        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 429)

    def test_quick_create_therapeutic_class_creates_lookup(self):
        response = self.client.post(
            reverse("items:quick_create_therapeutic_class"),
            {
                "code": "ABX",
                "name": "Antibiotik",
                "description": "Terapi infeksi",
            },
            secure=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(TherapeuticClass.objects.filter(code="ABX").exists())

    def test_quick_create_therapeutic_class_rejects_duplicate_name(self):
        TherapeuticClass.objects.create(code="ABX", name="Antibiotik")

        response = self.client.post(
            reverse("items:quick_create_therapeutic_class"),
            {
                "code": "ABX2",
                "name": "antibiotik",
                "description": "Duplikat",
            },
            secure=True,
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("Nama sudah digunakan", response.json()["error"])


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class ItemLookupRedirectSecurityTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(
            username="items_lookup_admin",
            email="items_lookup_admin@example.com",
            password="password12345",
        )
        self.client.force_login(self.user)

    def test_unit_create_ignores_external_next_on_get(self):
        response = self.client.get(
            reverse("items:unit_create"),
            {"next": "https://evil.example/phish"},
            secure=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'name="next"', html=False)
        self.assertContains(
            response,
            f'href="{reverse("items:item_create")}"',
            html=False,
        )
        self.assertNotContains(response, "https://evil.example/phish")

    def test_unit_create_redirects_to_local_same_host_next_path(self):
        response = self.client.post(
            reverse("items:unit_create"),
            {
                "code": "TAB",
                "name": "Tablet",
                "description": "Satuan tablet",
                "next": "https://testserver/items/create/?from=lookup",
            },
            secure=True,
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "/items/create/?from=lookup")

    def test_unit_create_falls_back_when_next_uses_external_host(self):
        response = self.client.post(
            reverse("items:unit_create"),
            {
                "code": "CAP",
                "name": "Capsule",
                "description": "Satuan kapsul",
                "next": "https://evil.example/items/create/",
            },
            secure=True,
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("items:item_create"))

    def test_unit_create_rejects_same_host_double_slash_path(self):
        response = self.client.post(
            reverse("items:unit_create"),
            {
                "code": "SYR",
                "name": "Syrup",
                "description": "Satuan sirup",
                "next": "https://testserver//evil.example/phish",
            },
            secure=True,
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("items:item_create"))

    def test_unit_create_ignores_same_host_backslash_path_on_get(self):
        response = self.client.get(
            reverse("items:unit_create"),
            {"next": r"https://testserver/\evil.example/phish"},
            secure=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'name="next"', html=False)
        self.assertContains(
            response,
            f'href="{reverse("items:item_create")}"',
            html=False,
        )
        self.assertNotContains(response, r"/\evil.example/phish")
