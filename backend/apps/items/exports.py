from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.core.xlsx_exports import escape_xlsx_formula

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _cell_value(value):
    return escape_xlsx_formula(value)


def export_items_excel(items, *, default_program=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Barang"

    ws.merge_cells("A1:H1")
    title = ws.cell(row=1, column=1, value=_cell_value("DAFTAR BARANG"))
    title.font = Font(bold=True, size=14)
    title.alignment = Alignment(horizontal="center")

    headers = [
        "Kode Barang",
        "Nama Barang",
        "Satuan",
        "Kategori",
        "Program",
        "Esensial",
        "Terapi Obat",
        "Minimum Stok",
    ]
    widths = [18, 36, 14, 18, 24, 14, 30, 16]
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=_cell_value(header))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, item in enumerate(items, start=4):
        if item.is_program_item:
            if item.program:
                program_label = f"[P] {item.program.code} - {item.program.name}"
            elif default_program:
                program_label = f"[P] {default_program.code} - {default_program.name}"
            else:
                program_label = "[P] DEFAULT"
        else:
            program_label = "-"

        therapeutic_names = ", ".join(
            therapeutic.name for therapeutic in sorted(item.therapeutic_classes.all(), key=lambda tc: tc.name)
        )

        row_values = [
            item.kode_barang,
            item.nama_barang,
            item.satuan.name,
            item.kategori.name,
            program_label,
            "Ya" if item.is_essential else "Tidak",
            therapeutic_names or "-",
            item.minimum_stock,
        ]
        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_cell_value(value))
            cell.border = THIN_BORDER
            if col_idx == 8:
                cell.alignment = Alignment(horizontal="right")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Daftar_Barang.xlsx"'
    wb.save(response)
    return response
